"""Ekstraksi statistik untuk paparan .pptx "Analisis Karhutla — Hotspot, Kompleks
Kebakaran, dan Luas Areal Kebakaran Kementerian Kehutanan".

Cakupan angka:
  1. Rekap hotspot NASA FIRMS 1 Januari 2026 s.d. tanggal berjalan, per provinsi
     lalu per wilayah kerja Balai Perhutanan Sosial (kolom wilker_bps).
  2. KPS dengan kompleks kebakaran terbanyak dalam 30 hari terakhir (menu
     "Kompleks Kebakaran" / ST-DBSCAN, preset sensitivitas "sedang").
  3. Luas areal kebakaran resmi Kementerian Kehutanan (tabel burned_area_summary),
     per provinsi dan per fungsi kawasan hutan.
  Ditambah: tren bulanan, kualitas deteksi, sebaran waktu, kepadatan hotspot per
  1.000 ha, dan KPS rawan kronis.

Titik hotspot dipetakan ke poligon KPS/Hutan Adat lewat spatial join ST_Covers
saat kueri (sama seperti menu Kompleks Kebakaran) -- tabel hotspot_polygon_relation
tidak dipakai karena isinya tertinggal jauh dari data observasi terkini.

READ-ONLY. Tidak menulis apa pun ke basis data.

Jalankan dari backend/:
    .venv/bin/python extract_hotspot_deck_stats.py <output.json>
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

from app.core.config import get_settings
from app.services.hotspot_cluster_service import HotspotClusterService
from app.services.postgres_store import PostgresStore

# ------------------------------------------------------------------ periode
TZ_WIB = timezone(timedelta(hours=7))
NOW = datetime.now(TZ_WIB)
YEAR_START = "2026-01-01"
NEXT_DAY = (NOW.date() + timedelta(days=1)).isoformat()  # batas atas eksklusif
WINDOW_DAYS = 30
WINDOW_START = NOW - timedelta(days=WINDOW_DAYS)

PENDAMPING_XLSX = Path(__file__).resolve().parents[1] / (
    "Master_Data_Pendamping_PS_2026_lengkap_180826.xlsx"
)


def rows(cur, sql, params=None):
    cur.execute(sql, params or ())
    return [dict(r) for r in cur.fetchall()]


def one(cur, sql, params=None):
    cur.execute(sql, params or ())
    r = cur.fetchone()
    return dict(r) if r else {}


def _num(text) -> float | None:
    if text is None:
        return None
    s = str(text).strip().replace(".", "").replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        v = float(s)
        return v if v > 0 else None
    except ValueError:
        return None


def _norm(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(name or "").lower())


# ------------------------------------------------------------------ base temp table
def build_hs_base(cur) -> None:
    """Satu temp table: setiap titik hotspot 2026 + poligon KPS/HA yang menaunginya
    (spatial join ST_Covers, satu poligon per titik) + fungsi kawasan hutan."""
    cur.execute("DROP TABLE IF EXISTS hs_base")
    cur.execute(
        """
        CREATE TEMP TABLE hs_base ON COMMIT PRESERVE ROWS AS
        SELECT
            o.id AS o_id,
            o.detected_at,
            extract(month from o.detected_at AT TIME ZONE 'Asia/Jakarta')::int AS mm,
            extract(hour  from o.detected_at AT TIME ZONE 'Asia/Jakarta')::int AS hh,
            lower(COALESCE(NULLIF(o.confidence,''),'(kosong)')) AS conf,
            COALESCE(NULLIF(o.source,''),'(kosong)') AS source,
            p.id AS pid,
            p.nama_prov,
            COALESCE(NULLIF(p.wilker_bps,''),'(tidak tercatat)') AS wilker_bps,
            p.nama_kab,
            p.lembaga,
            p.skema,
            p.no_sk,
            p.layer_key,
            p.luas_final,
            COALESCE(NULLIF(hk.kelompok,''),'(di luar kawasan)') AS kelompok
        FROM hotspot_observations o
        LEFT JOIN LATERAL (
            SELECT pp.id, pp.nama_prov, pp.wilker_bps, pp.nama_kab, pp.lembaga,
                   pp.skema, pp.no_sk, pp.layer_key, pp.luas_final
            FROM polygon_metadata pp
            WHERE pp.is_active
              AND pp.layer_key = o.layer_key
              AND ST_Covers(pp.geometry, o.geom)
            ORDER BY pp.id ASC
            LIMIT 1
        ) p ON TRUE
        LEFT JOIN hotspot_kawasan_hutan hk ON hk.hotspot_id = o.id
        WHERE o.detected_at >= %s AND o.detected_at < %s
        """,
        (YEAR_START, NEXT_DAY),
    )
    cur.execute("CREATE INDEX ON hs_base (mm)")
    cur.execute("CREATE INDEX ON hs_base (pid)")


# ------------------------------------------------------------------ 1. HOTSPOT
def hotspot_block(cur) -> dict:
    out: dict = {}
    out["hs_total"] = one(cur, "SELECT count(*) n FROM hs_base")["n"]
    out["hs_in_polygon"] = one(cur, "SELECT count(*) n FROM hs_base WHERE pid IS NOT NULL")["n"]

    out["hs_monthly"] = rows(cur, """
        SELECT mm AS m, count(*) n FROM hs_base GROUP BY 1 ORDER BY 1
    """)

    out["hs_by_confidence"] = rows(cur, """
        SELECT conf AS confidence, count(*) n FROM hs_base GROUP BY 1 ORDER BY 2 DESC
    """)
    out["hs_by_source"] = rows(cur, """
        SELECT source, count(*) n FROM hs_base GROUP BY 1 ORDER BY 2 DESC
    """)
    out["hs_by_hour"] = rows(cur, """
        SELECT hh AS h, count(*) n FROM hs_base GROUP BY 1 ORDER BY 1
    """)

    out["hs_by_prov"] = rows(cur, """
        SELECT nama_prov, count(*) n FROM hs_base
        WHERE pid IS NOT NULL GROUP BY 1 ORDER BY 2 DESC
    """)
    out["hs_by_prov_wilker"] = rows(cur, """
        SELECT nama_prov, wilker_bps, count(*) n FROM hs_base
        WHERE pid IS NOT NULL GROUP BY 1,2 ORDER BY 1, 3 DESC
    """)
    out["hs_by_wilker"] = rows(cur, """
        SELECT wilker_bps, count(*) n, count(DISTINCT nama_prov) n_prov FROM hs_base
        WHERE pid IS NOT NULL GROUP BY 1 ORDER BY 2 DESC
    """)
    out["kps_with_hotspot"] = one(cur, """
        SELECT count(DISTINCT pid) n FROM hs_base WHERE pid IS NOT NULL
    """)["n"]
    out["hs_by_kawasan"] = rows(cur, """
        SELECT kelompok, count(*) n FROM hs_base GROUP BY 1 ORDER BY 2 DESC
    """)
    out["hs_by_layer"] = rows(cur, """
        SELECT COALESCE(layer_key,'(di luar poligon)') layer_key, count(*) n
        FROM hs_base GROUP BY 1 ORDER BY 2 DESC
    """)
    return out


# ------------------------------------------------------------------ 2. KOMPLEKS
def kompleks_block(cur) -> dict:
    svc = HotspotClusterService()
    preset = {"eps_km": 2.0, "eps_hours": 48.0, "min_samples": 4, "location_eps_km": 1.0}
    res = svc.compute_clusters(
        start_at=WINDOW_START.astimezone(timezone.utc),
        end_at=NOW.astimezone(timezone.utc),
        eps_km=preset["eps_km"],
        eps_hours=preset["eps_hours"],
        min_samples=preset["min_samples"],
        location_eps_km=preset["location_eps_km"],
    )
    clusters = res.get("clusters", [])

    def _hours(c):
        a, b = c.get("first_detected_at"), c.get("last_detected_at")
        if not a or not b:
            return 0.0
        if isinstance(a, str):
            a = datetime.fromisoformat(a)
        if isinstance(b, str):
            b = datetime.fromisoformat(b)
        return max((b - a).total_seconds() / 3600.0, 0.0)

    durations = [_hours(c) for c in clusters]

    by_poly: dict[int, dict] = {}
    for c in clusters:
        dp = c.get("dominant_polygon")
        if not dp or dp.get("polygon_metadata_id") is None:
            continue
        pid = int(dp["polygon_metadata_id"])
        b = by_poly.setdefault(pid, {
            "polygon_metadata_id": pid, "name": dp.get("name"),
            "wilker_bps": dp.get("wilker_bps"), "province_name": dp.get("province_name"),
            "kompleks": 0, "titik": 0,
        })
        b["kompleks"] += 1
        b["titik"] += int(dp.get("hotspot_count") or 0)

    top = sorted(by_poly.values(), key=lambda x: (-x["kompleks"], -x["titik"]))[:15]
    ids = [t["polygon_metadata_id"] for t in top]
    meta = {}
    if ids:
        for m in rows(cur, """
            SELECT id, lembaga, nama_prov, nama_kab, skema, no_sk
            FROM polygon_metadata WHERE id = ANY(%s)
        """, (ids,)):
            meta[m["id"]] = m
    for t in top:
        m = meta.get(t["polygon_metadata_id"], {})
        t["lembaga"] = m.get("lembaga") or t.get("name")
        t["nama_prov"] = m.get("nama_prov") or t.get("province_name")
        t["nama_kab"] = m.get("nama_kab")
        t["skema"] = m.get("skema")
        t["no_sk"] = m.get("no_sk")

    hour_hist: dict[int, int] = {}
    for p in res.get("points", []):
        d = p.get("detected_at")
        if isinstance(d, str):
            d = datetime.fromisoformat(d)
        if d is not None:
            hh = d.astimezone(TZ_WIB).hour
            hour_hist[hh] = hour_hist.get(hh, 0) + 1

    stats = res.get("stats", {})
    return {
        "window_start": WINDOW_START.date().isoformat(),
        "window_end": NOW.date().isoformat(),
        "sensitivity": "sedang",
        "n_kompleks": res.get("count", 0),
        "clustered_hotspots": stats.get("clustered_hotspots", 0),
        "total_hotspots_in_range": stats.get("total_hotspots_in_range", 0),
        "unclustered_hotspots": stats.get("unclustered_hotspots", 0),
        "durasi_jam_rata": round(sum(durations) / len(durations), 1) if durations else 0.0,
        "durasi_jam_maks": round(max(durations), 1) if durations else 0.0,
        "kompleks_besar": sum(1 for c in clusters if int(c.get("hotspot_count") or 0) >= 10),
        "kps_terlibat": len(by_poly),
        "top_kps": top,
        "hour_hist": [{"h": h, "n": hour_hist.get(h, 0)} for h in range(24)],
    }


# ------------------------------------------------------------------ 3. LUAS KLHK
def klhk_block(cur) -> dict:
    out: dict = {}
    out["total"] = one(cur, """
        SELECT count(*) baris, count(DISTINCT polygon_metadata_id) kps,
               round(sum(burned_area_ha)::numeric,1) ha,
               min(make_date(year,month,1)) bln_awal,
               max(make_date(year,month,1)) bln_akhir
        FROM burned_area_summary
    """)
    out["monthly"] = rows(cur, """
        SELECT year, month, count(*) baris, round(sum(burned_area_ha)::numeric,1) ha
        FROM burned_area_summary GROUP BY 1,2 ORDER BY 1,2
    """)
    out["by_prov"] = rows(cur, """
        SELECT p.nama_prov, count(DISTINCT b.polygon_metadata_id) kps,
               count(*) baris, round(sum(b.burned_area_ha)::numeric,1) ha
        FROM burned_area_summary b
        JOIN polygon_metadata p ON p.id = b.polygon_metadata_id
        GROUP BY 1 ORDER BY 4 DESC
    """)
    out["sources"] = [r["source"] for r in rows(
        cur, "SELECT DISTINCT source FROM burned_area_summary")]
    try:
        out["by_kawasan"] = rows(cur, """
            SELECT COALESCE(NULLIF(lbl.kelompok, ''), NULLIF(bk.kelompok,''),
                            '(tidak terklasifikasi)') kelompok,
                   round(sum(bk.luas_ha)::numeric,1) ha
            FROM burned_kemenhut_kawasan_hutan bk
            LEFT JOIN ref_fungsi_kawasan_label lbl ON lbl.kode = bk.fungsikws
            GROUP BY 1 ORDER BY 2 DESC
        """)
    except Exception as exc:
        out["by_kawasan"] = []
        out["by_kawasan_error"] = str(exc)
    out["top_kps"] = rows(cur, """
        SELECT p.lembaga, p.nama_prov, p.nama_kab, p.skema,
               round(sum(b.burned_area_ha)::numeric,1) ha,
               array_agg(DISTINCT b.month ORDER BY b.month) bulan
        FROM burned_area_summary b
        JOIN polygon_metadata p ON p.id = b.polygon_metadata_id
        GROUP BY 1,2,3,4 ORDER BY 5 DESC LIMIT 12
    """)
    return out


# ------------------------------------------------------------------ TAMBAHAN
def extra_block(cur) -> dict:
    out: dict = {}
    raw = rows(cur, """
        SELECT pid, lembaga, nama_prov, nama_kab, skema, luas_final, count(*) n
        FROM hs_base WHERE pid IS NOT NULL
        GROUP BY 1,2,3,4,5,6
        HAVING count(*) >= 5
    """)
    dens = []
    for r in raw:
        luas = _num(r["luas_final"])
        if not luas or luas < 50:
            continue
        dens.append({
            "lembaga": r["lembaga"], "nama_prov": r["nama_prov"],
            "nama_kab": r["nama_kab"], "skema": r["skema"],
            "luas_ha": round(luas, 1), "n": r["n"],
            "per_1000ha": round(r["n"] / luas * 1000.0, 2),
        })
    dens.sort(key=lambda x: -x["per_1000ha"])
    out["density_top"] = dens[:12]

    out["chronic"] = rows(cur, """
        SELECT lembaga, nama_prov, nama_kab, skema,
               count(DISTINCT mm) bulan_kena, count(*) n,
               min(detected_at)::date::text d0, max(detected_at)::date::text d1
        FROM hs_base WHERE pid IS NOT NULL
        GROUP BY 1,2,3,4
        HAVING count(DISTINCT mm) >= 2
        ORDER BY 5 DESC, 6 DESC
        LIMIT 15
    """)
    out["chronic_count"] = one(cur, """
        SELECT count(*) n FROM (
          SELECT pid FROM hs_base WHERE pid IS NOT NULL
          GROUP BY 1 HAVING count(DISTINCT mm) >= 2
        ) t
    """)["n"]
    return out


# ------------------------------------------------------------------ PROFIL
def profil_block(cur) -> dict:
    out: dict = {}
    out["kps_total"] = one(
        cur, "SELECT count(*) n FROM polygon_metadata WHERE is_active")["n"]
    out["kps_by_layer"] = rows(cur, """
        SELECT layer_key, count(*) n FROM polygon_metadata
        WHERE is_active GROUP BY 1 ORDER BY 2 DESC
    """)
    out["kps_by_prov"] = rows(cur, """
        SELECT nama_prov, count(*) n FROM polygon_metadata
        WHERE is_active GROUP BY 1 ORDER BY 2 DESC
    """)
    out["n_prov"] = len(out["kps_by_prov"])
    out["n_wilker"] = one(cur, """
        SELECT count(DISTINCT NULLIF(wilker_bps,'')) n
        FROM polygon_metadata WHERE is_active
    """)["n"]
    return out


# ------------------------------------------------------------------ PENDAMPING
def pendamping_lookup() -> dict:
    if not PENDAMPING_XLSX.exists():
        return {}
    try:
        import openpyxl
    except ImportError:
        return {}
    wb = openpyxl.load_workbook(PENDAMPING_XLSX, read_only=True, data_only=True)
    if "KPS-Pendamping" not in wb.sheetnames:
        return {}
    ws = wb["KPS-Pendamping"]
    it = ws.iter_rows(values_only=True)
    header = [str(h or "").strip() for h in next(it)]
    idx = {name: header.index(name) for name in (
        "Nama Lembaga", "NAMA PENDAMPING", "Nama Balai", "Provinsi", "NO HP",
    ) if name in header}
    lut: dict[str, dict] = {}
    for row in it:
        lem = row[idx["Nama Lembaga"]] if "Nama Lembaga" in idx else None
        if not lem:
            continue
        rec = lut.setdefault(_norm(lem), {
            "pendamping": [], "hp": [], "balai": None, "prov": None})
        pdm = row[idx.get("NAMA PENDAMPING", -1)] if "NAMA PENDAMPING" in idx else None
        if pdm and str(pdm).strip():
            nm = re.sub(r"\s+", " ", str(pdm).strip())
            if nm not in rec["pendamping"]:
                rec["pendamping"].append(nm)
        if not rec["balai"] and "Nama Balai" in idx:
            rec["balai"] = row[idx["Nama Balai"]]
        if not rec["prov"] and "Provinsi" in idx:
            rec["prov"] = row[idx["Provinsi"]]
    return lut


def attach_pendamping(deck: dict) -> None:
    lut = pendamping_lookup()
    deck["pendamping_tersedia"] = bool(lut)
    if not lut:
        return
    prio: list[dict] = []
    seen: set[str] = set()
    for src in (deck["kompleks"]["top_kps"], deck["extra"]["chronic"],
                deck["extra"]["density_top"]):
        for r in src[:6]:
            nm = r.get("lembaga")
            if not nm or _norm(nm) in seen:
                continue
            seen.add(_norm(nm))
            rec = lut.get(_norm(nm))
            pdm = ", ".join(rec["pendamping"][:3]) if rec and rec["pendamping"] else "—"
            prio.append({
                "lembaga": nm,
                "nama_prov": r.get("nama_prov"),
                "pendamping": pdm,
                "balai": (rec or {}).get("balai") or "—",
            })
    deck["pendamping_prioritas"] = prio[:12]


# ------------------------------------------------------------------ MAIN
def main():
    s = get_settings()
    store = PostgresStore(s.database_url)
    deck: dict = {
        "generated_at": NOW.isoformat(),
        "periode": {"start": YEAR_START, "end": NOW.date().isoformat()},
        "window_days": WINDOW_DAYS,
    }
    with store.connection() as conn, conn.cursor() as cur:
        build_hs_base(cur)
        deck["profil"] = profil_block(cur)
        deck["hotspot"] = hotspot_block(cur)
        deck["kompleks"] = kompleks_block(cur)
        deck["klhk"] = klhk_block(cur)
        deck["extra"] = extra_block(cur)
        cur.execute("DROP TABLE IF EXISTS hs_base")
    attach_pendamping(deck)

    dest = sys.argv[1] if len(sys.argv) > 1 else "-"
    text = json.dumps(deck, ensure_ascii=False, indent=2, default=str)
    if dest == "-":
        print(text)
    else:
        Path(dest).write_text(text, encoding="utf-8")
        print(f"wrote {dest}  ({len(text):,} bytes)")


if __name__ == "__main__":
    main()
