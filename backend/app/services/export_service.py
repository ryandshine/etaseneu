from collections import Counter
from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.hotspot_categories import (
    CATEGORY_ORDER,
    confidence_category,
    frp_category,
)
from app.services.polygon_fields import (
    fungsi_kawasan,
    kelompok_kawasan,
    nama_kawasan_hutan,
    polygon_field,
    provinsi_name,
    sk_date,
    sk_number,
    skema_name,
)
from app.services.skema_stats import (
    SkemaProvinsiMatrix,
    build_skema_provinsi_matrix,
    count_per_skema,
)


WIB = timezone(timedelta(hours=7))

EXPORT_HEADERS = [
    "No",
    "Nama Wilayah",
    # Balai PS, Provinsi, & Skema disisipkan setelah "Nama Wilayah" -- posisi
    # kolom sebelumnya (No, Nama Wilayah) sengaja dipertahankan supaya konsumen
    # lama yang membaca dua kolom pertama tidak ikut bergeser.
    "No. SK",
    "Tgl SK",
    "Skema",
    "Balai PS",
    "Provinsi",
    "Kabupaten",
    "Satelit",
    "Tanggal Deteksi",
    "Latitude",
    "Longitude",
    "Confidence",
    "Kategori FRP",
    "FRP (MW)",
    "Brightness",
    # Atribusi fungsi kawasan hutan KLHK -- disisipkan di akhir supaya urutan
    # kolom lama tidak bergeser untuk konsumen yang membaca berdasar posisi.
    "Fungsi Kawasan Hutan",
    "Nama Kawasan",
    "Kelompok",
]

DATA_SHEET_TITLE = "Data Hotspot"
DASHBOARD_SHEET_TITLE = "Dashboard"
SKEMA_SHEET_TITLE = "Skema per Provinsi"
BURNED_AREA_SHEET_TITLE = "Luas Terbakar"

# Merah bata: burned area sengaja TIDAK memakai palet hijau/amber laporan
# hotspot -- dua metrik ini beda satuan (hektar vs jumlah titik) dan beda
# kesegaran data (bulanan vs 3 jam), jadi warnanya dibedakan supaya pembaca
# tidak keliru menjumlahkan keduanya.
_BURN = "9B2C2C"
_BURN_SOFT = "FDF0EE"

# ── Palet Warna Eksekutif Modern ─────────────────────────────────────────────
_INK = "1B3A2B"        # Deep Forest Green (Utama)
_EMERALD = "2D6A4F"    # Emerald Green (Sekunder)
_ACCENT = "E0862A"     # Warm Amber (Aksen Highlight)
_HEADER_TEXT = "FFFFFF"# Putih
_CARD_BG = "F4F8F5"    # Card background
_CARD_BORDER = "CFD8D2"# Card border
_BAND = "F8FAF9"       # Zebra striping lembut
_MUTED = "5A655F"      # Text muted
_RULE = "D8DDD9"       # Border tabel
_TOTAL_BG = "EBF2ED"   # Background baris total

_THIN_RULE = Side(style="thin", color=_RULE)
_DOUBLE_RULE = Side(style="double", color=_INK)
_CARD_SIDE = Side(style="thin", color=_CARD_BORDER)

_CELL_BORDER = Border(left=_THIN_RULE, right=_THIN_RULE, top=_THIN_RULE, bottom=_THIN_RULE)
_TOTAL_BORDER = Border(left=_THIN_RULE, right=_THIN_RULE, top=_THIN_RULE, bottom=_DOUBLE_RULE)
_CARD_BOX = Border(left=_CARD_SIDE, right=_CARD_SIDE, top=_CARD_SIDE, bottom=_CARD_SIDE)


def _wib_datetime(detected_at: object) -> datetime | None:
    if not detected_at:
        return None
    if isinstance(detected_at, datetime):
        if detected_at.tzinfo is None:
            return detected_at.replace(tzinfo=timezone.utc).astimezone(WIB)
        return detected_at.astimezone(WIB)
    s = str(detected_at).strip()
    try:
        parsed = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(WIB)
    except ValueError:
        pass
    for fmt in ("%d-%m-%Y %H:%M WIB", "%d-%m-%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(s, fmt)
            if "WIB" in s:
                return parsed.replace(tzinfo=WIB)
            return parsed.replace(tzinfo=timezone.utc).astimezone(WIB)
        except ValueError:
            continue
    return None


def _balai_ps(hotspot: dict) -> str:
    return polygon_field(hotspot, "WILKER_BPS") or "Tanpa Balai"


def build_export_rows(hotspots: list[dict]) -> list[dict]:
    rows: list[dict] = []
    for index, hotspot in enumerate(hotspots, start=1):
        wib = _wib_datetime(hotspot.get("detected_at"))
        detected_at_wib = (
            wib.strftime("%d-%m-%Y %H:%M WIB") if wib else str(hotspot.get("detected_at", ""))
        )

        rows.append(
            {
                "No": index,
                "Nama Wilayah": hotspot.get("layer_name", ""),
                "No. SK": sk_number(hotspot),
                "Tgl SK": sk_date(hotspot),
                "Skema": skema_name(hotspot),
                "Balai PS": _balai_ps(hotspot),
                "Provinsi": provinsi_name(hotspot),
                "Kabupaten": polygon_field(hotspot, "NAMA_KAB"),
                "Satelit": hotspot.get("source", ""),
                "Tanggal Deteksi": detected_at_wib,
                "Latitude": hotspot.get("latitude", ""),
                "Longitude": hotspot.get("longitude", ""),
                "Confidence": confidence_category(hotspot),
                "Kategori FRP": frp_category(hotspot),
                "FRP (MW)": hotspot.get("frp", ""),
                "Brightness": hotspot.get("brightness", ""),
                "Fungsi Kawasan Hutan": fungsi_kawasan(hotspot),
                "Nama Kawasan": nama_kawasan_hutan(hotspot),
                "Kelompok": kelompok_kawasan(hotspot),
            }
        )
    return rows


# ── Agregasi dashboard ─────────────────────────────────────────────────────

def build_dashboard_summary(hotspots: list[dict]) -> dict:
    """Ringkasan yang dipakai sheet Dashboard.

    Dipisah dari penulisan Excel supaya angkanya bisa diuji tanpa harus
    membongkar file .xlsx.
    """
    rows = build_export_rows(hotspots)
    total = len(rows)

    moments = [m for m in (_wib_datetime(h.get("detected_at")) for h in hotspots) if m]

    per_bulan: Counter[str] = Counter()
    for moment in moments:
        per_bulan[moment.strftime("%Y-%m")] += 1

    kps_counter: Counter[str] = Counter()
    kps_meta: dict[str, dict] = {}
    for row in rows:
        name = row["Nama Wilayah"]
        if name:
            kps_counter[name] += 1
            if name not in kps_meta:
                kps_meta[name] = {
                    "provinsi": row.get("Provinsi") or "Tanpa Provinsi",
                    "skema": row.get("Skema") or "Tanpa Skema",
                    "balai": row.get("Balai PS") or "Tanpa Balai",
                }

    top_kps_list = kps_counter.most_common(10)
    top_kps_detail = [
        {
            "rank": idx,
            "name": name,
            "provinsi": kps_meta[name]["provinsi"],
            "skema": kps_meta[name]["skema"],
            "balai": kps_meta[name]["balai"],
            "count": count,
            "percent": (count / total) if total > 0 else 0,
        }
        for idx, (name, count) in enumerate(top_kps_list, start=1)
    ]

    return {
        "total": total,
        "periode_awal": min(moments).strftime("%d-%m-%Y") if moments else "-",
        "periode_akhir": max(moments).strftime("%d-%m-%Y") if moments else "-",
        "jumlah_kps": len({row["Nama Wilayah"] for row in rows if row["Nama Wilayah"]}),
        "jumlah_balai": len({row["Balai PS"] for row in rows if row["Balai PS"]}),
        "jumlah_provinsi": len({row["Provinsi"] for row in rows if row["Provinsi"]}),
        "jumlah_skema": len({row["Skema"] for row in rows if row["Skema"]}),
        "per_balai": Counter(row["Balai PS"] for row in rows).most_common(),
        "per_provinsi": Counter(row["Provinsi"] for row in rows).most_common(),
        "per_skema": count_per_skema(hotspots),
        "skema_per_provinsi": build_skema_provinsi_matrix(hotspots),
        "per_satelit": Counter(row["Satelit"] for row in rows if row.get("Satelit")).most_common(),
        "per_confidence": [
            (level, sum(1 for row in rows if row["Confidence"] == level))
            for level in CATEGORY_ORDER
        ],
        "per_frp": [
            (level, sum(1 for row in rows if row["Kategori FRP"] == level))
            for level in CATEGORY_ORDER
        ],
        "per_bulan": sorted(per_bulan.items()),
        "top_kps": top_kps_list,
        "top_kps_detail": top_kps_detail,
    }


# ── Penulisan sheet ────────────────────────────────────────────────────────

def _write_data_sheet(sheet, rows: list[dict]) -> None:
    sheet.append(EXPORT_HEADERS)
    for row in rows:
        sheet.append([row[header] for header in EXPORT_HEADERS])

    header_fill = PatternFill("solid", fgColor=_INK)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color=_HEADER_TEXT, size=10)
        cell.alignment = Alignment(vertical="center", horizontal="left")
    sheet.row_dimensions[1].height = 22

    widths = [5, 34, 40, 12, 14, 22, 20, 18, 15, 20, 12, 12, 12, 12, 10, 12, 26, 24, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"
    if rows:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(EXPORT_HEADERS))}{len(rows) + 1}"


def _title_banner(sheet, summary: dict) -> None:
    sheet.merge_cells("A1:O1")
    sheet["A1"] = "ETA SENEU  •  DASHBOARD MONITORING HOTSPOT KAWASAN"
    sheet["A1"].font = Font(name="Calibri", bold=True, size=14, color=_HEADER_TEXT)
    sheet["A1"].fill = PatternFill("solid", fgColor=_INK)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 26

    sheet.merge_cells("A2:O2")
    sheet["A2"] = "Sistem Peringatan Dini & Rekapitulasi Intersep Titik Panas Poligon Perhutanan Sosial (KPS)"
    sheet["A2"].font = Font(name="Calibri", size=9.5, color="E2EFE7", italic=True)
    sheet["A2"].fill = PatternFill("solid", fgColor=_EMERALD)
    sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[2].height = 20

    sheet.merge_cells("A3:O3")
    meta_text = (
        f"Periode Data: {summary['periode_awal']} s.d. {summary['periode_akhir']}   •   "
        f"Total Hotspot: {summary['total']:,} Titik   •   "
        f"KPS Terdampak: {summary['jumlah_kps']:,} Unit SK   •   "
        f"Cakupan: {summary['jumlah_balai']} Balai PS / {summary['jumlah_provinsi']} Provinsi"
    )
    sheet["A3"] = meta_text
    sheet["A3"].font = Font(name="Calibri", bold=True, size=9, color=_INK)
    sheet["A3"].fill = PatternFill("solid", fgColor=_TOTAL_BG)
    sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[3].height = 20
    sheet.row_dimensions[4].height = 8


def _write_kpi_cards(sheet, summary: dict) -> None:
    cards = [
        ("TOTAL HOTSPOT", summary["total"], "Titik Panas Terdeteksi", 1, 3, _ACCENT),
        ("KPS TERDAMPAK", summary["jumlah_kps"], "Unit SK Kawasan Hutan", 4, 6, _EMERALD),
        ("BALAI PS", summary["jumlah_balai"], "Wilayah Kerja Balai PS", 7, 9, _EMERALD),
        ("PROVINSI", summary["jumlah_provinsi"], "Provinsi Teridentifikasi", 10, 12, _EMERALD),
        ("SKEMA PS", summary["jumlah_skema"], "Skema Izin Terdaftar", 13, 15, _EMERALD),
    ]
    card_fill = PatternFill("solid", fgColor=_CARD_BG)

    for label, value, subtitle, c_start, c_end, accent_color in cards:
        accent_fill = PatternFill("solid", fgColor=accent_color)
        sheet.cell(row=5, column=c_start, value=label)
        sheet.cell(row=6, column=c_start, value=value)
        sheet.cell(row=7, column=c_start, value=subtitle)

        for r in range(5, 9):
            for c in range(c_start, c_end + 1):
                cell = sheet.cell(row=r, column=c)
                cell.fill = accent_fill if r == 8 else card_fill
                cell.border = _CARD_BOX

        lbl_cell = sheet.cell(row=5, column=c_start)
        lbl_cell.font = Font(name="Calibri", bold=True, size=8.5, color=_MUTED)
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")

        val_cell = sheet.cell(row=6, column=c_start)
        val_color = _ACCENT if accent_color == _ACCENT else _INK
        val_cell.font = Font(name="Calibri", bold=True, size=18, color=val_color)
        val_cell.number_format = "#,##0"
        val_cell.alignment = Alignment(horizontal="center", vertical="center")

        sub_cell = sheet.cell(row=7, column=c_start)
        sub_cell.font = Font(name="Calibri", italic=True, size=8, color=_MUTED)
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")

        sheet.merge_cells(start_row=5, start_column=c_start, end_row=5, end_column=c_end)
        sheet.merge_cells(start_row=6, start_column=c_start, end_row=6, end_column=c_end)
        sheet.merge_cells(start_row=7, start_column=c_start, end_row=7, end_column=c_end)
        sheet.merge_cells(start_row=8, start_column=c_start, end_row=8, end_column=c_end)

    sheet.row_dimensions[5].height = 16
    sheet.row_dimensions[6].height = 26
    sheet.row_dimensions[7].height = 14
    sheet.row_dimensions[8].height = 4
    sheet.row_dimensions[9].height = 10


def _write_section_header(sheet, row: int, col_start: int, col_end: int, title: str) -> None:
    sheet.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = sheet.cell(row=row, column=col_start, value=title)
    cell.font = Font(name="Calibri", bold=True, size=10, color=_HEADER_TEXT)
    cell.fill = PatternFill("solid", fgColor=_INK)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(col_start, col_end + 1):
        c_cell = sheet.cell(row=row, column=c)
        c_cell.border = _CELL_BORDER
        c_cell.fill = PatternFill("solid", fgColor=_INK)
    sheet.row_dimensions[row].height = 22


def _attach_bar_chart(sheet, anchor: str, title: str, header_row: int, row_count: int,
                      chart_type: str = "bar", color: str = _INK, data_col: int = 1,
                      height: float = 7.5, width: float = 16.5,
                      y_axis_title: str = "Jumlah Titik Panas") -> None:
    if row_count <= 0:
        return
    chart = BarChart()
    chart.type = chart_type
    chart.title = title
    chart.style = 10
    chart.legend = None
    chart.height = height
    chart.width = width

    if chart_type == "bar":
        chart.x_axis.axPos = "l"
        chart.x_axis.tickLblPos = "nextTo"
        chart.x_axis.title = None
        chart.y_axis.axPos = "b"
        chart.y_axis.tickLblPos = "nextTo"
        chart.y_axis.title = y_axis_title
        chart.y_axis.majorGridlines = ChartLines()
    else:
        chart.x_axis.axPos = "b"
        chart.x_axis.tickLblPos = "nextTo"
        chart.x_axis.title = None
        chart.y_axis.axPos = "l"
        chart.y_axis.tickLblPos = "nextTo"
        chart.y_axis.title = y_axis_title
        chart.y_axis.majorGridlines = ChartLines()

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showCatName = False
    chart.dataLabels.showSerName = False
    chart.dataLabels.showPercent = False
    chart.dataLabels.showLegendKey = False

    value_col = data_col + 1
    data = Reference(sheet, min_col=value_col, min_row=header_row, max_row=header_row + row_count)
    categories = Reference(sheet, min_col=data_col, min_row=header_row + 1, max_row=header_row + row_count)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    cat_col_letter = get_column_letter(data_col)
    cat_range = f"'{sheet.title}'!${cat_col_letter}${header_row + 1}:${cat_col_letter}${header_row + row_count}"
    if chart.series:
        chart.series[0].cat = AxDataSource(strRef=StrRef(f=cat_range))
        chart.series[0].graphicalProperties.solidFill = color

    sheet.add_chart(chart, anchor)


def _attach_line_chart(sheet, anchor: str, title: str, header_row: int, row_count: int,
                       color: str = _INK, accent_color: str = _ACCENT, data_col: int = 1,
                       height: float = 7.5, width: float = 16.5) -> None:
    if row_count <= 0:
        return
    chart = LineChart()
    chart.title = title
    chart.style = 13
    chart.height = height
    chart.width = width
    chart.legend = None

    chart.x_axis.axPos = "b"
    chart.x_axis.tickLblPos = "nextTo"
    chart.x_axis.title = "Bulan"

    chart.y_axis.axPos = "l"
    chart.y_axis.tickLblPos = "nextTo"
    chart.y_axis.title = "Jumlah Titik Panas"
    chart.y_axis.majorGridlines = ChartLines()

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showCatName = False
    chart.dataLabels.showSerName = False
    chart.dataLabels.showPercent = False
    chart.dataLabels.showLegendKey = False

    value_col = data_col + 1
    data = Reference(sheet, min_col=value_col, min_row=header_row, max_row=header_row + row_count)
    categories = Reference(sheet, min_col=data_col, min_row=header_row + 1, max_row=header_row + row_count)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    cat_col_letter = get_column_letter(data_col)
    cat_range = f"'{sheet.title}'!${cat_col_letter}${header_row + 1}:${cat_col_letter}${header_row + row_count}"
    if chart.series:
        chart.series[0].cat = AxDataSource(strRef=StrRef(f=cat_range))
        chart.series[0].graphicalProperties.line.solidFill = color
        chart.series[0].graphicalProperties.line.width = 25000
        chart.series[0].marker.symbol = "circle"
        chart.series[0].marker.size = 6
        chart.series[0].marker.graphicalProperties.solidFill = accent_color
        chart.series[0].marker.graphicalProperties.line.solidFill = accent_color

    sheet.add_chart(chart, anchor)


def _write_chart_data(sheet, row: int, col: int, headers: tuple[str, str],
                      data: list[tuple[str, int]]) -> int:
    header_row = row
    sheet.cell(row=header_row, column=col, value=headers[0])
    sheet.cell(row=header_row, column=col + 1, value=headers[1])
    for index, (label, count) in enumerate(data, start=1):
        sheet.cell(row=header_row + index, column=col, value=label)
        sheet.cell(row=header_row + index, column=col + 1, value=count)
    return header_row


def _write_mini_summary_table(sheet, start_row: int, start_col: int, title: str,
                              data: list[tuple[str, int]], total: int) -> int:
    headers = [title, "Jumlah", "%"]
    header_fill = PatternFill("solid", fgColor=_EMERALD)
    for idx, h in enumerate(headers):
        cell = sheet.cell(row=start_row, column=start_col + idx, value=h)
        cell.fill = header_fill
        cell.font = Font(name="Calibri", bold=True, color=_HEADER_TEXT, size=8.5)
        cell.alignment = Alignment(horizontal="center" if idx > 0 else "left", vertical="center")
        cell.border = _CELL_BORDER
    sheet.row_dimensions[start_row].height = 18

    current_row = start_row + 1
    for idx, (label, count) in enumerate(data):
        pct = (count / total) if total > 0 else 0
        r_cell = sheet.cell(row=current_row, column=start_col, value=label)
        r_cell.font = Font(name="Calibri", size=8.5)
        r_cell.alignment = Alignment(horizontal="left", vertical="center")
        r_cell.border = _CELL_BORDER
        if idx % 2 == 1:
            r_cell.fill = PatternFill("solid", fgColor=_BAND)

        c_cell = sheet.cell(row=current_row, column=start_col + 1, value=count)
        c_cell.font = Font(name="Calibri", size=8.5)
        c_cell.number_format = "#,##0"
        c_cell.alignment = Alignment(horizontal="right", vertical="center")
        c_cell.border = _CELL_BORDER
        if idx % 2 == 1:
            c_cell.fill = PatternFill("solid", fgColor=_BAND)

        p_cell = sheet.cell(row=current_row, column=start_col + 2, value=pct)
        p_cell.font = Font(name="Calibri", size=8.5)
        p_cell.number_format = "0.0%"
        p_cell.alignment = Alignment(horizontal="right", vertical="center")
        p_cell.border = _CELL_BORDER
        if idx % 2 == 1:
            p_cell.fill = PatternFill("solid", fgColor=_BAND)

        sheet.row_dimensions[current_row].height = 16
        current_row += 1

    sum_counts = sum(count for _, count in data)
    sum_pct = (sum_counts / total) if total > 0 else 0
    t_lbl = sheet.cell(row=current_row, column=start_col, value="Total")
    t_lbl.font = Font(name="Calibri", bold=True, size=8.5, color=_INK)
    t_lbl.fill = PatternFill("solid", fgColor=_TOTAL_BG)
    t_lbl.border = _TOTAL_BORDER

    t_val = sheet.cell(row=current_row, column=start_col + 1, value=sum_counts)
    t_val.font = Font(name="Calibri", bold=True, size=8.5, color=_INK)
    t_val.number_format = "#,##0"
    t_val.alignment = Alignment(horizontal="right", vertical="center")
    t_val.fill = PatternFill("solid", fgColor=_TOTAL_BG)
    t_val.border = _TOTAL_BORDER

    t_pct = sheet.cell(row=current_row, column=start_col + 2, value=sum_pct)
    t_pct.font = Font(name="Calibri", bold=True, size=8.5, color=_INK)
    t_pct.number_format = "0.0%"
    t_pct.alignment = Alignment(horizontal="right", vertical="center")
    t_pct.fill = PatternFill("solid", fgColor=_TOTAL_BG)
    t_pct.border = _TOTAL_BORDER

    sheet.row_dimensions[current_row].height = 18
    return current_row


def _write_top_kps_table(sheet, start_row: int, start_col: int, top_kps_detail: list[dict], total: int) -> int:
    headers = [
        ("No", "center"),
        ("Nama Kawasan / KPS", "left"),
        ("Provinsi", "left"),
        ("Skema", "center"),
        ("Balai PS", "left"),
        ("Jumlah", "right"),
        ("% Total", "right"),
    ]
    header_fill = PatternFill("solid", fgColor=_EMERALD)

    for idx, (title, align) in enumerate(headers):
        cell = sheet.cell(row=start_row, column=start_col + idx, value=title)
        cell.fill = header_fill
        cell.font = Font(name="Calibri", bold=True, color=_HEADER_TEXT, size=8.5)
        cell.alignment = Alignment(horizontal="center" if idx < 5 else align, vertical="center")
        cell.border = _CELL_BORDER

    sheet.row_dimensions[start_row].height = 18
    current_row = start_row + 1

    for row_idx, item in enumerate(top_kps_detail):
        fill_color = _BAND if row_idx % 2 == 1 else "FFFFFF"
        row_fill = PatternFill("solid", fgColor=fill_color)

        c_no = sheet.cell(row=current_row, column=start_col, value=f"#{item['rank']}")
        c_no.alignment = Alignment(horizontal="center", vertical="center")
        c_no.font = Font(name="Calibri", bold=True, size=8, color=_MUTED)

        c_name = sheet.cell(row=current_row, column=start_col + 1, value=item["name"])
        c_name.alignment = Alignment(horizontal="left", vertical="center")
        c_name.font = Font(name="Calibri", bold=True, size=8.5, color=_INK)

        c_prov = sheet.cell(row=current_row, column=start_col + 2, value=item["provinsi"])
        c_prov.alignment = Alignment(horizontal="left", vertical="center")
        c_prov.font = Font(name="Calibri", size=8)

        c_skema = sheet.cell(row=current_row, column=start_col + 3, value=item["skema"])
        c_skema.alignment = Alignment(horizontal="center", vertical="center")
        c_skema.font = Font(name="Calibri", bold=True, size=8, color=_EMERALD)

        c_balai = sheet.cell(row=current_row, column=start_col + 4, value=item["balai"])
        c_balai.alignment = Alignment(horizontal="left", vertical="center")
        c_balai.font = Font(name="Calibri", size=8)

        c_count = sheet.cell(row=current_row, column=start_col + 5, value=item["count"])
        c_count.alignment = Alignment(horizontal="right", vertical="center")
        c_count.font = Font(name="Calibri", bold=True, size=8.5, color=_ACCENT)
        c_count.number_format = "#,##0"

        c_pct = sheet.cell(row=current_row, column=start_col + 6, value=item["percent"])
        c_pct.alignment = Alignment(horizontal="right", vertical="center")
        c_pct.font = Font(name="Calibri", size=8)
        c_pct.number_format = "0.0%"

        for offset in range(7):
            cell = sheet.cell(row=current_row, column=start_col + offset)
            cell.border = _CELL_BORDER
            cell.fill = row_fill

        sheet.row_dimensions[current_row].height = 17
        current_row += 1

    if not top_kps_detail:
        empty_cell = sheet.cell(row=current_row, column=start_col, value="Tidak ada data KPS terdampak")
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
        empty_cell.font = Font(name="Calibri", italic=True, size=8.5, color=_MUTED)
        sheet.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=start_col + 6)
        for offset in range(7):
            sheet.cell(row=current_row, column=start_col + offset).border = _CELL_BORDER
        sheet.row_dimensions[current_row].height = 20
        current_row += 1

    sum_top_counts = sum(item["count"] for item in top_kps_detail)
    sum_top_pct = (sum_top_counts / total) if total > 0 else 0

    sheet.cell(row=current_row, column=start_col, value="Total 10 KPS Teratas")
    sheet.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=start_col + 4)
    lbl = sheet.cell(row=current_row, column=start_col)
    lbl.alignment = Alignment(horizontal="right", vertical="center")
    lbl.font = Font(name="Calibri", bold=True, size=8.5, color=_INK)

    v_cell = sheet.cell(row=current_row, column=start_col + 5, value=sum_top_counts)
    v_cell.alignment = Alignment(horizontal="right", vertical="center")
    v_cell.font = Font(name="Calibri", bold=True, size=8.5, color=_INK)
    v_cell.number_format = "#,##0"

    p_cell = sheet.cell(row=current_row, column=start_col + 6, value=sum_top_pct)
    p_cell.alignment = Alignment(horizontal="right", vertical="center")
    p_cell.font = Font(name="Calibri", bold=True, size=8.5, color=_INK)
    p_cell.number_format = "0.0%"

    for offset in range(7):
        cell = sheet.cell(row=current_row, column=start_col + offset)
        cell.border = _TOTAL_BORDER
        cell.fill = PatternFill("solid", fgColor=_TOTAL_BG)

    sheet.row_dimensions[current_row].height = 18
    return current_row


def _write_dashboard_sheet(sheet, summary: dict) -> None:
    total = summary["total"]

    # 1. Header Banner & Metadata Card
    _title_banner(sheet, summary)

    # 2. 5 KPI Scorecard Cards (Baris 5 - 8)
    _write_kpi_cards(sheet, summary)

    # 3. Card Baris 1: Distribusi Wilayah Balai & Skema PS (Baris 10 - 27)
    _write_section_header(sheet, 10, 1, 7, "1. SEBARAN HOTSPOT PER BALAI PS")
    _write_section_header(sheet, 10, 9, 15, "2. SEBARAN HOTSPOT PER SKEMA PERHUTANAN SOSIAL")

    # 4. Card Baris 2: Tren Waktu & Distribusi Provinsi (Baris 29 - 46)
    _write_section_header(sheet, 29, 1, 7, "3. TREN DETEKSI HOTSPOT BULANAN")
    _write_section_header(sheet, 29, 9, 15, "4. DISTRIBUSI SEBARAN PER PROVINSI")

    # 5. Card Baris 3: Parameter Risiko & 10 KPS Prioritas (Baris 48 - 63)
    _write_section_header(sheet, 48, 1, 7, "5. PARAMETER TINGKAT RISIKO & SENSOR SATELIT")
    _write_section_header(sheet, 48, 9, 15, "6. 10 UNIT KPS PRIORITAS PENANGANAN (HOTSPOT TERBANYAK)")

    # Tabel Parameter Risiko di Kolom A:C
    _write_mini_summary_table(sheet, 50, 1, "Tingkat Kepercayaan", summary["per_confidence"], total)
    _write_mini_summary_table(sheet, 56, 1, "Kategori FRP", summary["per_frp"], total)

    # Tabel Satelit Sensor di Kolom E:G
    satelit_data = summary["per_satelit"] if summary["per_satelit"] else [("Semua Sensor", total)]
    _write_mini_summary_table(sheet, 50, 5, "Satelit Sensor", satelit_data, total)

    # Tabel 10 KPS Prioritas di Kolom I:O
    _write_top_kps_table(sheet, 50, 9, summary.get("top_kps_detail", []), total)

    # 6. Pemasangan Chart & Penulisan Data Staging
    charts = [
        ("Hotspot per Balai PS", ("Balai PS", "Jumlah"), summary["per_balai"], "bar", _INK, "A11"),
        ("Hotspot per Skema PS", ("Skema", "Jumlah"), summary["per_skema"], "col", _EMERALD, "I11"),
        ("Tren Bulanan", ("Bulan", "Jumlah"), summary["per_bulan"], "line", _INK, "A30"),
        ("Hotspot per Provinsi", ("Provinsi", "Jumlah"), summary["per_provinsi"], "bar", _ACCENT, "I30"),
    ]

    staging_row = 70
    sheet.cell(row=staging_row - 2, column=1, value="Data pendukung chart (di luar area cetak dashboard)")
    sheet.cell(row=staging_row - 2, column=1).font = Font(name="Calibri", italic=True, size=8, color=_MUTED)

    for title, headers, data, chart_kind, color, anchor in charts:
        header_row = _write_chart_data(sheet, staging_row, 1, headers, data)
        staging_row = header_row + len(data) + 3

        if chart_kind == "line":
            _attach_line_chart(sheet, anchor, title, header_row, len(data), color,
                               height=7.2, width=16.2)
        elif chart_kind == "col":
            _attach_bar_chart(sheet, anchor, title, header_row, len(data), chart_type="col", color=color,
                              height=7.2, width=16.2)
        else:
            _attach_bar_chart(sheet, anchor, title, header_row, len(data), chart_type="bar", color=color,
                              height=7.2, width=16.2)

    # Pengaturan Lebar Kolom Grid Modular (Simetris 2 Panel A..G dan I..O)
    col_widths = {
        "A": 14, "B": 10, "C": 8,
        "D": 3,
        "E": 16, "F": 10, "G": 8,
        "H": 4,
        "I": 6, "J": 28, "K": 16, "L": 12, "M": 18, "N": 12, "O": 9
    }
    for col_letter, width in col_widths.items():
        sheet.column_dimensions[col_letter].width = width

    # Print & View setup
    sheet.print_area = "A1:O63"
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_margins.left = 0.3
    sheet.page_margins.right = 0.3
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.3
    sheet.page_margins.header = 0.2
    sheet.page_margins.footer = 0.2

    sheet.views.sheetView[0].showGridLines = True


def _write_skema_provinsi_sheet(sheet, matrix: SkemaProvinsiMatrix) -> None:
    """Tabel silang provinsi (baris) x skema (kolom) beserta baris/kolom total."""
    sheet["A1"] = "REKAP HOTSPOT PER SKEMA PERHUTANAN SOSIAL PER PROVINSI"
    sheet["A1"].font = Font(bold=True, size=14, color=_INK)
    sheet["A2"] = (
        "Jumlah titik panas pada setiap kombinasi provinsi dan skema izin perhutanan sosial. "
        "Kolom diurutkan dari skema dengan titik panas terbanyak."
    )
    sheet["A2"].font = Font(size=10, color=_MUTED, italic=True)

    header_row = 4
    headers = ("Provinsi", *matrix.skema, "Total")
    header_fill = PatternFill("solid", fgColor=_INK)
    for offset, header in enumerate(headers):
        cell = sheet.cell(row=header_row, column=1 + offset, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color=_HEADER_TEXT, size=10)
        cell.border = _CELL_BORDER
        cell.alignment = Alignment(
            vertical="center", horizontal="left" if offset == 0 else "center", wrap_text=True
        )
    sheet.row_dimensions[header_row].height = 28

    for index, row in enumerate(matrix.rows):
        row_number = header_row + 1 + index
        values = (row.provinsi, *row.counts, row.total)
        for offset, value in enumerate(values):
            cell = sheet.cell(row=row_number, column=1 + offset, value=value)
            cell.border = _CELL_BORDER
            if offset:
                cell.alignment = Alignment(horizontal="center")
            if index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=_BAND)
        sheet.cell(row=row_number, column=len(headers)).font = Font(bold=True, color=_INK)

    total_row = header_row + len(matrix.rows) + 1
    for offset, value in enumerate(("Total", *matrix.totals, matrix.grand_total)):
        cell = sheet.cell(row=total_row, column=1 + offset, value=value)
        cell.font = Font(bold=True, color=_HEADER_TEXT, size=10)
        cell.fill = PatternFill("solid", fgColor=_ACCENT)
        cell.border = _CELL_BORDER
        if offset:
            cell.alignment = Alignment(horizontal="center")

    sheet.column_dimensions["A"].width = 30
    for offset in range(1, len(headers)):
        sheet.column_dimensions[get_column_letter(1 + offset)].width = 16
    sheet.freeze_panes = f"B{header_row + 1}"


def _write_burned_area_sheet(sheet, report: dict) -> None:
    """Lampiran luas bekas terbakar: rekap per skema + rincian per KPS.

    Angka hektar di sini TIDAK sebanding dengan jumlah titik hotspot di sheet
    lain -- sumbernya rekap resmi KLHK yang terbit tidak dengan jadwal tetap,
    bukan deteksi near-real-time. Peringatan itu ditulis di sheet supaya
    pembaca yang cuma membuka lampiran ini tidak salah menyimpulkan.
    """
    sheet["A1"] = "REKAP LUAS BEKAS TERBAKAR PADA KAWASAN PERHUTANAN SOSIAL"
    sheet["A1"].font = Font(bold=True, size=14, color=_BURN)
    sheet["A2"] = (
        "Luas area terbakar hasil overlay batas KPS dengan poligon resmi KLHK "
        "\"Areal Kebakaran Hutan dan Lahan\" (akurasi H/M -- terverifikasi hotspot, sesuai "
        "kriteria KLHK). Luas dihitung sekali per kawasan walau terbakar berulang, sehingga "
        "tidak dapat dijumlahkan langsung dengan jumlah titik hotspot."
    )
    sheet["A2"].font = Font(size=9.5, color=_MUTED, italic=True)
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells("A2:G2")
    sheet.row_dimensions[2].height = 28

    rows = report.get("rows") or []
    if not rows:
        sheet["A4"] = "Belum ada data luas terbakar untuk filter laporan ini."
        sheet["A4"].font = Font(size=11, color=_MUTED, italic=True)
        sheet.column_dimensions["A"].width = 60
        return

    total_cell = sheet["A4"]
    total_cell.value = (
        f"TOTAL {report['total_ha']:,.1f} Ha  •  {report['kps_count']} KPS terdampak"
    )
    total_cell.font = Font(bold=True, size=11, color=_HEADER_TEXT)
    total_cell.fill = PatternFill("solid", fgColor=_BURN)
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells("A4:G4")
    sheet.row_dimensions[4].height = 22

    # ── Rekap per skema ──────────────────────────────────────────────────
    skema_header_row = 6
    sheet.cell(row=skema_header_row - 1, column=1, value="A. REKAPITULASI PER SKEMA").font = Font(
        bold=True, size=10, color=_INK
    )
    for offset, title in enumerate(("Skema", "Luas Terbakar (Ha)", "Jumlah KPS", "% Total")):
        cell = sheet.cell(row=skema_header_row, column=1 + offset, value=title)
        cell.fill = PatternFill("solid", fgColor=_BURN)
        cell.font = Font(bold=True, color=_HEADER_TEXT, size=9)
        cell.border = _CELL_BORDER
        cell.alignment = Alignment(horizontal="left" if offset == 0 else "center", vertical="center")

    total_ha = float(report["total_ha"]) or 1.0
    for index, item in enumerate(report.get("by_skema") or []):
        row_number = skema_header_row + 1 + index
        values = (
            item["skema"],
            round(float(item["total_ha"]), 1),
            int(item["kps_count"]),
            float(item["total_ha"]) / total_ha,
        )
        for offset, value in enumerate(values):
            cell = sheet.cell(row=row_number, column=1 + offset, value=value)
            cell.border = _CELL_BORDER
            cell.font = Font(size=9)
            if offset:
                cell.alignment = Alignment(horizontal="center")
            if index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=_BURN_SOFT)
        sheet.cell(row=row_number, column=2).number_format = "#,##0.0"
        sheet.cell(row=row_number, column=4).number_format = "0.0%"

    # Chart batang luas terbakar per skema, ditumpuk di kanan tabel rekap --
    # sebelumnya sheet ini cuma berisi tabel angka, tanpa representasi visual
    # sama sekali (beda dari sheet Dashboard yang penuh chart). Anchor-nya di
    # kolom F (tabel rekap cuma sampai D) supaya tidak menimpa tabel di
    # kirinya, tapi tingginya (~7cm, ~13 baris) tetap harus diperhitungkan
    # supaya tabel rincian KPS di bawahnya tidak dimulai sebelum chart ini
    # selesai -- kolom F-I dipakai keduanya (Luas/Bulan/Periode/Keterangan).
    _CHART_ROW_SPAN = 15
    chart_anchor_row = skema_header_row - 1
    _attach_bar_chart(
        sheet,
        f"F{chart_anchor_row}",
        "Luas Terbakar per Skema (Ha)",
        skema_header_row,
        len(report.get("by_skema") or []),
        chart_type="col",
        color=_BURN,
        data_col=1,
        height=7.0,
        width=13.0,
        y_axis_title="Luas Terbakar (Ha)",
    )

    # ── Rincian per KPS ──────────────────────────────────────────────────
    detail_start = max(
        skema_header_row + len(report.get("by_skema") or []) + 3,
        chart_anchor_row + _CHART_ROW_SPAN,
    )
    sheet.cell(row=detail_start - 1, column=1, value="B. RINCIAN PER KAWASAN (KPS)").font = Font(
        bold=True, size=10, color=_INK
    )

    detail_headers = (
        "No",
        "Nama Kawasan / KPS",
        "Skema",
        "Provinsi",
        "Balai PS",
        "Luas Terbakar (Ha)",
        "Bulan Terbakar",
        "Periode Terakhir",
        "Keterangan",
    )
    for offset, title in enumerate(detail_headers):
        cell = sheet.cell(row=detail_start, column=1 + offset, value=title)
        cell.fill = PatternFill("solid", fgColor=_BURN)
        cell.font = Font(bold=True, color=_HEADER_TEXT, size=9)
        cell.border = _CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[detail_start].height = 26

    ranked = sorted(rows, key=lambda item: item["burned_area_ha"], reverse=True)
    for index, item in enumerate(ranked):
        row_number = detail_start + 1 + index
        values = (
            index + 1,
            item["lembaga"],
            item["skema"],
            item["nama_prov"],
            item["wilker_bps"],
            round(float(item["burned_area_ha"]), 1),
            item["burned_months"],
            item["latest_period"],
            # Dibedakan eksplisit: baris "perkiraan" luasnya di bawah resolusi
            # piksel citra, jadi angkanya benar tapi lokasinya tidak presisi.
            "Perkiraan lokasi" if item["is_estimated"] else "Terpetakan",
        )
        for offset, value in enumerate(values):
            cell = sheet.cell(row=row_number, column=1 + offset, value=value)
            cell.border = _CELL_BORDER
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="left" if offset in (1, 3, 4) else "center")
            if index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=_BURN_SOFT)
        sheet.cell(row=row_number, column=2).font = Font(size=9, bold=True, color=_INK)
        ha_cell = sheet.cell(row=row_number, column=6)
        ha_cell.number_format = "#,##0.0"
        ha_cell.font = Font(size=9, bold=True, color=_BURN)

    total_row = detail_start + len(ranked) + 1
    sheet.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color=_HEADER_TEXT, size=9)
    sheet.cell(row=total_row, column=6, value=round(float(report["total_ha"]), 1)).number_format = "#,##0.0"
    for offset in range(len(detail_headers)):
        cell = sheet.cell(row=total_row, column=1 + offset)
        cell.fill = PatternFill("solid", fgColor=_BURN)
        cell.font = Font(bold=True, color=_HEADER_TEXT, size=9)
        cell.border = _TOTAL_BORDER
        cell.alignment = Alignment(horizontal="center")

    widths = (5, 38, 10, 20, 22, 17, 14, 15, 16)
    for offset, width in enumerate(widths):
        sheet.column_dimensions[get_column_letter(1 + offset)].width = width
    sheet.column_dimensions["A"].width = 5
    sheet.freeze_panes = f"A{detail_start + 1}"


def build_excel_file(hotspots: list[dict], burned_area_report: dict | None = None) -> bytes:
    workbook = Workbook()

    summary = build_dashboard_summary(hotspots)

    # 1. Sheet Dashboard sebagai lembar kerja utama
    dashboard_sheet = workbook.active
    dashboard_sheet.title = DASHBOARD_SHEET_TITLE
    _write_dashboard_sheet(dashboard_sheet, summary)

    # 2. Sheet Skema per Provinsi
    skema_sheet = workbook.create_sheet(SKEMA_SHEET_TITLE)
    _write_skema_provinsi_sheet(skema_sheet, summary["skema_per_provinsi"])

    # 3. Sheet Luas Terbakar (opsional -- hanya kalau pemanggil menyertakan
    #    datanya, supaya pemanggil lain yang tidak butuh tidak berubah).
    if burned_area_report is not None:
        burned_sheet = workbook.create_sheet(BURNED_AREA_SHEET_TITLE)
        _write_burned_area_sheet(burned_sheet, burned_area_report)

    # 4. Sheet Data Hotspot
    data_sheet = workbook.create_sheet(DATA_SHEET_TITLE)
    _write_data_sheet(data_sheet, build_export_rows(hotspots))

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
