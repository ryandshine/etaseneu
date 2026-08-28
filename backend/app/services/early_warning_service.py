"""Layanan Analisis Peringatan Dini (Early Warning) & Rekapitulasi Kebakaran KPS."""

from datetime import datetime
import io
import math
from typing import Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.config import get_settings
from app.services.postgres_store import PostgresStore


def compute_ftri_score(
    h_today: int,
    h_yesterday: int,
    h_7d: int,
    h_aug: int,
    h_total_2026: int,
    total_burned_ha: float,
    burn_frequency: int,
    luas_sk: float | None,
    h_today_strict_reburn: int = 0,
) -> float:
    """Hitung Skor FTRI (Fire Threat & Recurrence Index) 0 - 100."""
    h_now = (h_today * 10) + (h_today_strict_reburn * 5) + (h_yesterday * 3) + (h_7d * 0.5)
    urgency_score = min(50.0, 10.0 * math.log1p(h_now))
    
    severity_score = min(30.0, (5.0 * math.log1p(total_burned_ha)) + (burn_frequency * 2.0))
    
    luas_sk_val = float(luas_sk) if luas_sk and float(luas_sk) > 0 else 100.0
    ratio = min(1.0, total_burned_ha / luas_sk_val)
    ratio_score = ratio * 20.0
    
    if total_burned_ha == 0:
        score_trend = min(30.0, 6.0 * math.log1p(h_7d * 2 + h_aug))
        score_total = min(20.0, 4.0 * math.log1p(h_total_2026))
        return round(urgency_score + score_trend + score_total, 1)
        
    return round(urgency_score + severity_score + ratio_score, 1)


class EarlyWarningService:
    def __init__(self, store: PostgresStore | None = None):
        self.store = store or PostgresStore(get_settings().database_url)

    def get_summary_metrics(self) -> dict[str, Any]:
        """Ambil metrik agregat makro status kebakaran KPS."""
        with self.store.connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    WITH polygon_burned AS (
                        SELECT 
                            polygon_metadata_id,
                            SUM(burned_area_ha) as total_burned_ha,
                            ST_Union(geometry) FILTER (WHERE geometry IS NOT NULL) as burned_geom
                        FROM burned_area_summary
                        WHERE burned_area_ha > 0
                        GROUP BY polygon_metadata_id
                    ),
                    polygon_hotspots AS (
                        SELECT 
                            p.id as polygon_id,
                            COUNT(h.id) FILTER (WHERE h.detected_at >= DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta')) as h_today,
                            COUNT(h.id) FILTER (
                                WHERE h.detected_at >= DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta')
                                  AND b.burned_geom IS NOT NULL
                                  AND ST_Contains(b.burned_geom, h.geom)
                            ) as h_today_strict_reburn,
                            COUNT(h.id) FILTER (WHERE h.detected_at >= DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta' - INTERVAL '1 day') AND h.detected_at < DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta')) as h_yesterday,
                            COUNT(h.id) FILTER (WHERE h.detected_at >= DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta' - INTERVAL '7 days')) as h_7d,
                            COUNT(h.id) FILTER (WHERE h.detected_at >= DATE_TRUNC('month', NOW() AT TIME ZONE 'Asia/Jakarta')) as h_month
                        FROM polygon_metadata p
                        JOIN polygon_burned b ON p.id = b.polygon_metadata_id
                        JOIN hotspot_observations h ON ST_Contains(p.geometry, h.geom)
                        WHERE p.is_active = TRUE
                        GROUP BY p.id
                    )
                    SELECT 
                        COUNT(b.polygon_metadata_id) as total_burned_polygons,
                        COALESCE(SUM(b.total_burned_ha), 0) as total_burned_ha,
                        COUNT(*) FILTER (WHERE COALESCE(h.h_today, 0) > 0) as burned_active_today,
                        COUNT(*) FILTER (WHERE COALESCE(h.h_today_strict_reburn, 0) > 0) as burned_strict_reburn_kps_today,
                        COALESCE(SUM(h.h_today_strict_reburn), 0) as total_strict_reburn_hotspots_today,
                        COALESCE(SUM(h.h_today - h.h_today_strict_reburn), 0) as total_expanding_hotspots_today,
                        COUNT(*) FILTER (WHERE COALESCE(h.h_today, 0) = 0) as burned_clear_today,
                        COUNT(*) FILTER (WHERE COALESCE(h.h_today, 0) = 0 AND COALESCE(h.h_yesterday, 0) > 0) as burned_active_yesterday,
                        COUNT(*) FILTER (WHERE COALESCE(h.h_today, 0) = 0 AND COALESCE(h.h_yesterday, 0) = 0 AND COALESCE(h.h_7d, 0) > 0) as burned_active_7d,
                        COUNT(*) FILTER (WHERE COALESCE(h.h_today, 0) = 0 AND COALESCE(h.h_yesterday, 0) = 0 AND COALESCE(h.h_7d, 0) = 0 AND COALESCE(h.h_month, 0) = 0) as burned_padam_total
                    FROM polygon_metadata p
                    JOIN polygon_burned b ON p.id = b.polygon_metadata_id
                    LEFT JOIN polygon_hotspots h ON p.id = h.polygon_id
                    WHERE p.is_active = TRUE
                    """
                )
                burned_stats = cur.fetchone() or {}

                cur.execute(
                    """
                    WITH ew_polygons AS (
                        SELECT 
                            p.id,
                            COUNT(h.id) FILTER (WHERE h.detected_at >= DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta')) as h_today,
                            COUNT(h.id) FILTER (WHERE h.detected_at >= DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta' - INTERVAL '1 day') AND h.detected_at < DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta')) as h_yesterday,
                            COUNT(h.id) FILTER (WHERE h.detected_at >= DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta' - INTERVAL '7 days')) as h_7d,
                            COUNT(h.id) FILTER (WHERE h.detected_at >= DATE_TRUNC('month', NOW() AT TIME ZONE 'Asia/Jakarta')) as h_month,
                            COUNT(h.id) as h_total
                        FROM polygon_metadata p
                        JOIN hotspot_observations h ON ST_Contains(p.geometry, h.geom)
                        LEFT JOIN (
                            SELECT polygon_metadata_id
                            FROM burned_area_summary
                            WHERE year = EXTRACT(YEAR FROM NOW()) AND burned_area_ha > 0
                            GROUP BY polygon_metadata_id
                        ) b ON p.id = b.polygon_metadata_id
                        WHERE p.is_active = TRUE
                          AND b.polygon_metadata_id IS NULL
                          AND h.detected_at >= DATE_TRUNC('year', NOW() AT TIME ZONE 'Asia/Jakarta')
                        GROUP BY p.id
                    )
                    SELECT 
                        COUNT(*) as ew_total_kps,
                        COUNT(*) FILTER (WHERE h_today > 0) as ew_today_kps,
                        COUNT(*) FILTER (WHERE h_today = 0 AND h_yesterday > 0) as ew_yesterday_kps,
                        COUNT(*) FILTER (WHERE h_today = 0 AND h_yesterday = 0 AND h_7d > 0) as ew_7d_kps,
                        COALESCE(SUM(h_today), 0) as ew_today_hotspots,
                        COALESCE(SUM(h_month), 0) as ew_month_hotspots,
                        COALESCE(SUM(h_total), 0) as ew_year_hotspots
                    FROM ew_polygons
                    """
                )
                ew_stats = cur.fetchone() or {}

        return {
            "burned_area_stats": {
                "total_polygons": int(burned_stats.get("total_burned_polygons") or 0),
                "total_burned_ha": float(burned_stats.get("total_burned_ha") or 0.0),
                "active_today": int(burned_stats.get("burned_active_today") or 0),
                "strict_reburn_kps_today": int(burned_stats.get("burned_strict_reburn_kps_today") or 0),
                "strict_reburn_hotspots_today": int(burned_stats.get("total_strict_reburn_hotspots_today") or 0),
                "expanding_hotspots_today": int(burned_stats.get("total_expanding_hotspots_today") or 0),
                "clear_today": int(burned_stats.get("burned_clear_today") or 0),
                "active_yesterday": int(burned_stats.get("burned_active_yesterday") or 0),
                "active_7d": int(burned_stats.get("burned_active_7d") or 0),
                "padam_total": int(burned_stats.get("burned_padam_total") or 0),
            },
            "early_warning_stats": {
                "total_kps": int(ew_stats.get("ew_total_kps") or 0),
                "active_today": int(ew_stats.get("ew_today_kps") or 0),
                "active_yesterday": int(ew_stats.get("ew_yesterday_kps") or 0),
                "active_7d": int(ew_stats.get("ew_7d_kps") or 0),
                "today_hotspots": int(ew_stats.get("ew_today_hotspots") or 0),
                "month_hotspots": int(ew_stats.get("ew_month_hotspots") or 0),
                "year_hotspots": int(ew_stats.get("ew_year_hotspots") or 0),
            },
            "updated_at": datetime.now().isoformat(),
        }

    def get_kps_analysis_list(
        self,
        *,
        category: str = "burned_active_today",
        province: str | None = None,
        skema: str | None = None,
        search: str | None = None,
        limit: int = 1500,
    ) -> list[dict[str, Any]]:
        """Ambil daftar KPS dengan klasifikasi zona perambatan ilmiah."""
        with self.store.connection() as conn:
            with conn.cursor() as cur:
                is_burned_filter = category in [
                    "all_burned",
                    "burned_active_today",
                    "burned_clear_today",
                    "burned_active_yesterday",
                    "burned_active_7d",
                    "burned_padam_total",
                ]

                base_query = """
                    WITH polygon_burned AS (
                        SELECT 
                            polygon_metadata_id,
                            SUM(burned_area_ha) as total_burned_ha,
                            COUNT(DISTINCT (year, month)) as burn_frequency,
                            MAX(make_date(year, month, 1)) as latest_burned_month,
                            ST_Union(geometry) FILTER (WHERE geometry IS NOT NULL) as burned_geom
                        FROM burned_area_summary
                        WHERE burned_area_ha > 0
                        GROUP BY polygon_metadata_id
                    ),
                    polygon_hotspots AS (
                        SELECT 
                            p.id as polygon_id,
                            COUNT(h.id) FILTER (WHERE h.detected_at >= DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta')) as h_today,
                            COUNT(h.id) FILTER (
                                WHERE h.detected_at >= DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta')
                                  AND b.burned_geom IS NOT NULL
                                  AND ST_Contains(b.burned_geom, h.geom)
                            ) as h_today_strict_reburn,
                            MIN(
                                CASE 
                                    WHEN h.detected_at >= DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta')
                                         AND b.burned_geom IS NOT NULL
                                         AND NOT ST_Contains(b.burned_geom, h.geom)
                                    THEN ST_Distance(h.geom::geography, b.burned_geom::geography) / 1000.0
                                    ELSE NULL
                                END
                            ) as min_distance_km_today,
                            MAX(
                                CASE 
                                    WHEN h.detected_at >= DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta')
                                         AND b.burned_geom IS NOT NULL
                                         AND NOT ST_Contains(b.burned_geom, h.geom)
                                    THEN ST_Distance(h.geom::geography, b.burned_geom::geography) / 1000.0
                                    ELSE NULL
                                END
                            ) as max_distance_km_today,
                            AVG(
                                CASE 
                                    WHEN h.detected_at >= DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta')
                                         AND b.burned_geom IS NOT NULL
                                         AND NOT ST_Contains(b.burned_geom, h.geom)
                                    THEN ST_Distance(h.geom::geography, b.burned_geom::geography) / 1000.0
                                    ELSE NULL
                                END
                            ) as avg_distance_km_today,
                            COUNT(h.id) FILTER (WHERE h.detected_at >= DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta' - INTERVAL '1 day') AND h.detected_at < DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta')) as h_yesterday,
                            COUNT(h.id) FILTER (WHERE h.detected_at >= DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta' - INTERVAL '7 days')) as h_7d,
                            COUNT(h.id) FILTER (
                                WHERE h.detected_at >= DATE_TRUNC('day', NOW() AT TIME ZONE 'Asia/Jakarta' - INTERVAL '7 days')
                                  AND b.burned_geom IS NOT NULL
                                  AND ST_Contains(b.burned_geom, h.geom)
                            ) as h_7d_strict_reburn,
                            COUNT(h.id) FILTER (WHERE h.detected_at >= DATE_TRUNC('month', NOW() AT TIME ZONE 'Asia/Jakarta')) as h_month,
                            COUNT(h.id) FILTER (WHERE h.detected_at >= DATE_TRUNC('year', NOW() AT TIME ZONE 'Asia/Jakarta')) as h_year,
                            MAX(h.detected_at) as latest_hotspot_at
                        FROM polygon_metadata p
                        LEFT JOIN polygon_burned b ON p.id = b.polygon_metadata_id
                        JOIN hotspot_observations h ON ST_Contains(p.geometry, h.geom)
                        WHERE p.is_active = TRUE
                        GROUP BY p.id
                    )
                    SELECT 
                        p.id,
                        p.lembaga,
                        p.nama_desa,
                        p.nama_kec,
                        p.nama_kab,
                        p.nama_prov,
                        p.skema,
                        p.luas_sk,
                        p.no_sk,
                        COALESCE(b.total_burned_ha, 0) as total_burned_ha,
                        COALESCE(b.burn_frequency, 0) as burn_frequency,
                        b.latest_burned_month,
                        COALESCE(h.h_today, 0) as h_today,
                        COALESCE(h.h_today_strict_reburn, 0) as h_today_strict_reburn,
                        h.min_distance_km_today,
                        h.max_distance_km_today,
                        h.avg_distance_km_today,
                        COALESCE(h.h_yesterday, 0) as h_yesterday,
                        COALESCE(h.h_7d, 0) as h_7d,
                        COALESCE(h.h_7d_strict_reburn, 0) as h_7d_strict_reburn,
                        COALESCE(h.h_month, 0) as h_month,
                        COALESCE(h.h_year, 0) as h_year,
                        h.latest_hotspot_at
                    FROM polygon_metadata p
                    LEFT JOIN polygon_burned b ON p.id = b.polygon_metadata_id
                    LEFT JOIN polygon_hotspots h ON p.id = h.polygon_id
                    WHERE p.is_active = TRUE
                """

                clauses = []
                params: list[Any] = []

                if is_burned_filter:
                    clauses.append("b.total_burned_ha > 0")
                    if category == "burned_active_today":
                        clauses.append("COALESCE(h.h_today, 0) > 0")
                    elif category == "burned_clear_today":
                        clauses.append("COALESCE(h.h_today, 0) = 0")
                    elif category == "burned_active_yesterday":
                        clauses.append("COALESCE(h.h_today, 0) = 0 AND COALESCE(h.h_yesterday, 0) > 0")
                    elif category == "burned_active_7d":
                        clauses.append("COALESCE(h.h_today, 0) = 0 AND COALESCE(h.h_yesterday, 0) = 0 AND COALESCE(h.h_7d, 0) > 0")
                    elif category == "burned_padam_total":
                        clauses.append("COALESCE(h.h_today, 0) = 0 AND COALESCE(h.h_month, 0) = 0")
                elif category.startswith("early_warning"):
                    clauses.append("(b.total_burned_ha IS NULL OR b.total_burned_ha = 0)")
                    clauses.append("COALESCE(h.h_year, 0) > 0")
                    if category == "early_warning_today":
                        clauses.append("COALESCE(h.h_today, 0) > 0")
                    elif category == "early_warning_yesterday":
                        clauses.append("COALESCE(h.h_today, 0) = 0 AND COALESCE(h.h_yesterday, 0) > 0")
                    elif category == "early_warning_7d":
                        clauses.append("COALESCE(h.h_today, 0) = 0 AND COALESCE(h.h_yesterday, 0) = 0 AND COALESCE(h.h_7d, 0) > 0")

                if province:
                    clauses.append("p.nama_prov = %s")
                    params.append(province)
                if skema:
                    clauses.append("p.skema = %s")
                    params.append(skema)
                if search:
                    clauses.append("(p.lembaga ILIKE %s OR p.nama_kab ILIKE %s OR p.nama_kec ILIKE %s OR p.nama_desa ILIKE %s)")
                    term = f"%{search}%"
                    params.extend([term, term, term, term])

                where_sql = f" AND {' AND '.join(clauses)}" if clauses else ""
                
                order_sql = """
                    ORDER BY 
                        (COALESCE(h.h_today, 0) > 0) DESC,
                        COALESCE(h.h_today, 0) DESC,
                        (COALESCE(h.h_yesterday, 0) > 0) DESC,
                        COALESCE(h.h_yesterday, 0) DESC,
                        (COALESCE(h.h_7d, 0) > 0) DESC,
                        COALESCE(h.h_7d, 0) DESC,
                        COALESCE(b.total_burned_ha, 0) DESC
                    LIMIT %s
                """
                params.append(limit)

                cur.execute(base_query + where_sql + order_sql, params)
                rows = cur.fetchall()

        results = []
        for r in rows:
            h_today = r["h_today"]
            h_strict = r["h_today_strict_reburn"]
            h_expand = max(0, h_today - h_strict)
            
            h_7d = r["h_7d"]
            h_7d_strict = r["h_7d_strict_reburn"]
            h_7d_expand = max(0, h_7d - h_7d_strict)

            min_dist = round(float(r["min_distance_km_today"]), 2) if r["min_distance_km_today"] is not None else None
            max_dist = round(float(r["max_distance_km_today"]), 2) if r["max_distance_km_today"] is not None else None
            avg_dist = round(float(r["avg_distance_km_today"]), 2) if r["avg_distance_km_today"] is not None else None

            # Klasifikasi Zona Ilmiah
            if r["total_burned_ha"] > 0:
                if h_today > 0:
                    if h_strict > 0 and h_expand == 0:
                        propagation_zone = "Strict Re-burn (Bara Bekas)"
                        zone_code = "strict"
                    elif h_strict > 0 and h_expand > 0:
                        propagation_zone = "Kombinasi Bara & Merambat"
                        zone_code = "combo"
                    elif min_dist is not None and min_dist <= 1.0:
                        propagation_zone = "Zona 1: Merambat Langsung (≤1.0 km)"
                        zone_code = "zone1"
                    elif min_dist is not None and min_dist <= 3.0:
                        propagation_zone = "Zona 2: Loncatan Bara / Spotting (1-3 km)"
                        zone_code = "zone2"
                    else:
                        propagation_zone = "Zona 3: Titik Bakar Mandiri (>3.0 km)"
                        zone_code = "zone3"
                else:
                    propagation_zone = "-"
                    zone_code = "none"
            else:
                propagation_zone = "Titik Baru 2026 (Belum Ada Rekap)"
                zone_code = "new_2026"

            ftri = compute_ftri_score(
                h_today=h_today,
                h_yesterday=r["h_yesterday"],
                h_7d=h_7d,
                h_aug=r["h_month"],
                h_total_2026=r["h_year"],
                total_burned_ha=float(r["total_burned_ha"] or 0),
                burn_frequency=r["burn_frequency"],
                luas_sk=float(r["luas_sk"]) if r["luas_sk"] else None,
                h_today_strict_reburn=h_strict,
            )

            # Label status & detail perambatan
            if r["total_burned_ha"] > 0:
                if h_today > 0:
                    dist_text = f" ({min_dist}-{max_dist} km)" if min_dist is not None and max_dist is not None and max_dist > min_dist else f" ({min_dist} km)" if min_dist is not None else ""
                    if h_strict > 0 and h_expand > 0:
                        status_badge = f"🔴 Bara Bekas ({h_strict}) & Blok Baru ({h_expand}{dist_text})"
                    elif h_strict > 0:
                        status_badge = f"🔴 Strict Re-burn ({h_strict} di Bekas Terbakar)"
                    elif zone_code == "zone1":
                        status_badge = f"🔴 Zona 1: Merambat Langsung ({h_expand} Baru{dist_text})"
                    elif zone_code == "zone2":
                        status_badge = f"🟠 Zona 2: Loncatan Bara ({h_expand} Baru{dist_text})"
                    else:
                        status_badge = f"🟡 Zona 3: Titik Bakar Mandiri ({h_expand} Baru{dist_text})"
                elif r["h_yesterday"] > 0:
                    status_badge = "⚠️ Reda Kemarin"
                elif h_7d > 0:
                    status_badge = "🟡 Reda 7 Hari"
                elif r["h_month"] > 0:
                    status_badge = "⚪ Reda di Awal Bulan"
                else:
                    status_badge = "🟢 Padam / 0 Hotspot"
            else:
                if h_today > 0:
                    status_badge = "🔴 P1: Titik Baru Hari Ini"
                elif r["h_yesterday"] > 0:
                    status_badge = "🟠 P2: Titik Baru Kemarin"
                elif h_7d > 0:
                    status_badge = "🟡 P3: Titik Baru 7 Hari"
                else:
                    status_badge = "🟢 Terpantau Jan-Juli"

            results.append({
                "id": r["id"],
                "lembaga": r["lembaga"],
                "nama_desa": r["nama_desa"],
                "nama_kec": r["nama_kec"],
                "nama_kab": r["nama_kab"],
                "nama_prov": r["nama_prov"],
                "skema": r["skema"],
                "luas_sk": float(r["luas_sk"]) if r["luas_sk"] else None,
                "no_sk": r["no_sk"],
                "total_burned_ha": float(r["total_burned_ha"] or 0),
                "burn_frequency": r["burn_frequency"],
                "latest_burned_month": r["latest_burned_month"].strftime("%B %Y") if r["latest_burned_month"] else None,
                "hotspots_today": h_today,
                "hotspots_today_strict_reburn": h_strict,
                "hotspots_today_expanding": h_expand,
                "min_distance_km": min_dist,
                "max_distance_km": max_dist,
                "avg_distance_km": avg_dist,
                "propagation_zone": propagation_zone,
                "zone_code": zone_code,
                "hotspots_yesterday": r["h_yesterday"],
                "hotspots_7d": h_7d,
                "hotspots_7d_strict_reburn": h_7d_strict,
                "hotspots_7d_expanding": h_7d_expand,
                "hotspots_month": r["h_month"],
                "hotspots_year": r["h_year"],
                "latest_hotspot_at": r["latest_hotspot_at"].isoformat() if r["latest_hotspot_at"] else None,
                "ftri_score": ftri,
                "status_label": status_badge,
            })
        return results

    def build_excel_export(self, category: str = "all") -> bytes:
        """Bangun file Excel ekspor terstruktur langsung dari memori."""
        wb = openpyxl.Workbook()
        font_title = Font(name="Calibri", size=15, bold=True, color="1B365D")
        font_subtitle = Font(name="Calibri", size=10, italic=True, color="555555")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_regular = Font(name="Calibri", size=10)

        fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        fill_red_light = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        fill_orange_light = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        fill_yellow_light = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")

        border_thin = Border(
            left=Side(style="thin", color="D3D3D3"), right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"), bottom=Side(style="thin", color="D3D3D3")
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        ws = wb.active
        ws.title = "Rekap Analisis Kebakaran"
        ws.views.sheetView[0].showGridLines = True

        ws["A1"] = "REKAPITULASI ANALISIS KEBAKARAN & PERINGATAN DINI KPS"
        ws["A1"].font = font_title
        ws["A2"] = f"Dihasilkan pada: {datetime.now().strftime('%d %B %Y %H:%M WIB')} | Kategori: {category.upper()}"
        ws["A2"].font = font_subtitle

        headers = [
            "No", "ID", "Nama KPS / Lembaga", "Skema", "Desa", "Kecamatan", "Kabupaten", "Provinsi",
            "Luas SK (ha)", "Luas Terbakar KLHK (ha)", "Frekuensi Terbakar",
            "Hotspot Hari Ini (Total)", "Tepat di Bekas Terbakar (Strict Re-burn)", "Perembetan Blok Baru",
            "Jarak Perambatan Min (KM)", "Jarak Perambatan Max (KM)", "Jarak Perambatan Rerata (KM)",
            "Klasifikasi Zona Ilmiah",
            "Hotspot Kemarin", "Hotspot 7 Hari", "Hotspot Bulan Ini", "Total Hotspot 2026",
            "Deteksi Terakhir", "Skor FTRI", "Status & Rincian Mekanisme"
        ]

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        items = self.get_kps_analysis_list(category=category, limit=2000)
        for row_idx, r in enumerate(items, 5):
            dt_str = r["latest_hotspot_at"][:16].replace("T", " ") if r["latest_hotspot_at"] else "-"
            row_values = [
                row_idx - 4, r["id"], r["lembaga"], r["skema"], r["nama_desa"] or "-", r["nama_kec"] or "-",
                r["nama_kab"] or "-", r["nama_prov"] or "-", r["luas_sk"], r["total_burned_ha"], r["burn_frequency"],
                r["hotspots_today"], r["hotspots_today_strict_reburn"], r["hotspots_today_expanding"],
                r["min_distance_km"], r["max_distance_km"], r["avg_distance_km"],
                r["propagation_zone"],
                r["hotspots_yesterday"], r["hotspots_7d"], r["hotspots_month"], r["hotspots_year"],
                dt_str, r["ftri_score"], r["status_label"]
            ]
            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_regular
                cell.border = border_thin
                if r["hotspots_today"] > 0:
                    if col_idx in [1, 12, 13, 14, 18, 24, 25]:
                        if r["zone_code"] == "zone1" or r["zone_code"] == "strict" or r["zone_code"] == "combo":
                            cell.fill = fill_red_light
                        elif r["zone_code"] == "zone2":
                            cell.fill = fill_orange_light
                        else:
                            cell.fill = fill_yellow_light
                elif r["hotspots_yesterday"] > 0:
                    if col_idx in [1, 19, 24, 25]:
                        cell.fill = fill_orange_light

                if col_idx in [1, 2, 4, 11, 18, 23, 25]:
                    cell.alignment = align_center
                elif col_idx in [9, 10, 12, 13, 14, 15, 16, 17, 19, 20, 21, 22, 24]:
                    cell.alignment = align_right
                    if col_idx in [9, 10, 15, 16, 17, 24]:
                        cell.number_format = "#,##0.00"
                    else:
                        cell.number_format = "#,##0"
                else:
                    cell.alignment = align_left

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len and cell.row > 2:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
