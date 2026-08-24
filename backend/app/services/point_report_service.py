"""Laporan hasil pencocokan titik unggahan ke KPS: Excel dan PDF.

Gaya visual (warna, header, band) sengaja mengikuti export_service supaya
berkas dari dua fitur berbeda tidak terasa berasal dari dua aplikasi berbeda.
"""

from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.hotspot_categories import confidence_category, frp_category
from app.services.point_match_service import MatchOutcome

WIB = timezone(timedelta(hours=7))

DATA_SHEET_TITLE = "Data Titik"
DASHBOARD_SHEET_TITLE = "Dashboard"

_INK = "1B3A2B"
_ACCENT = "E0862A"
_HEADER_TEXT = "FFFFFF"
_BAND = "EEF3EF"
_MUTED = "6B726D"
_OUTSIDE = "B4453C"

_THIN_RULE = Side(style="thin", color="D8DEDA")
_CELL_BORDER = Border(left=_THIN_RULE, right=_THIN_RULE, top=_THIN_RULE, bottom=_THIN_RULE)

# Kolom hasil pencocokan. Metadata asli pengguna ditambahkan SETELAH ini,
# supaya kolom tetap sama posisinya apa pun isi berkas yang diunggah.
BASE_HEADERS = [
    "No",
    "Latitude",
    "Longitude",
    "Status",
    "Kategori Confidence",
    "Kategori FRP",
    "KPS (Lembaga)",
    "Balai PS",
    "Provinsi",
    "Kabupaten",
    "Kecamatan",
    "Desa",
    "Skema",
    "No. SK",
    "Tgl SK",
]

_KPS_FIELD_BY_HEADER = {
    "KPS (Lembaga)": "lembaga",
    "Balai PS": "wilker_bps",
    "Provinsi": "nama_prov",
    "Kabupaten": "nama_kab",
    "Kecamatan": "nama_kec",
    "Desa": "nama_desa",
    "Skema": "skema",
    "No. SK": "no_sk",
    "Tgl SK": "tgl_sk",
}

STATUS_INSIDE = "Masuk KPS"
STATUS_OUTSIDE = "Di luar KPS"

# Nama kolom confidence/FRP tidak seragam antar sumber berkas (mis. "confidence"
# vs "Confidence", "frp" vs "FRP (MW)"), jadi dicari tanpa peduli besar-kecil
# huruf/spasi alih-alih mengharuskan nama kolom persis.
_CONFIDENCE_KEYS = ("confidence", "conf")
_FRP_KEYS = ("frp", "frp_mw", "frpmw")


def _find_property(properties: dict[str, Any], candidates: tuple[str, ...]) -> Any:
    lowered = {str(key).strip().lower().replace(" ", "").replace("_", ""): value for key, value in properties.items()}
    for candidate in candidates:
        key = candidate.replace("_", "")
        if key in lowered and lowered[key] not in (None, ""):
            return lowered[key]
    return None


def _confidence_and_frp_category(properties: dict[str, Any]) -> tuple[str, str]:
    """Kategori confidence & FRP dari properti asli titik, kalau ada.

    Memakai ambang batas yang sama dengan dashboard utama (hotspot_categories.py)
    supaya konsisten. Kalau berkas yang diunggah tidak punya kolom confidence/FRP
    sama sekali, hasilnya "-" -- BUKAN dianggap "Rendah", karena "tidak ada data"
    dan "nilainya rendah" adalah dua hal berbeda.
    """
    confidence_raw = _find_property(properties, _CONFIDENCE_KEYS)
    frp_raw = _find_property(properties, _FRP_KEYS)
    confidence = confidence_category({"confidence": confidence_raw}) if confidence_raw is not None else "-"
    frp = frp_category({"frp": frp_raw}) if frp_raw is not None else "-"
    return confidence, frp


def _stringify(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def build_report_rows(outcome: MatchOutcome) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for index, point in enumerate(outcome.points, start=1):
        kps = point.kps or {}
        kategori_confidence, kategori_frp = _confidence_and_frp_category(point.properties)
        row: list[Any] = [
            index,
            round(point.latitude, 6),
            round(point.longitude, 6),
            STATUS_INSIDE if point.inside_kps else STATUS_OUTSIDE,
            kategori_confidence,
            kategori_frp,
        ]
        for header in BASE_HEADERS[6:]:
            row.append(_stringify(kps.get(_KPS_FIELD_BY_HEADER[header])))
        for column in outcome.property_columns:
            row.append(_stringify(point.properties.get(column)))
        rows.append(row)
    return rows


def _write_data_sheet(sheet, outcome: MatchOutcome) -> None:
    headers = [*BASE_HEADERS, *outcome.property_columns]
    sheet.append(headers)

    for row in build_report_rows(outcome):
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor=_INK)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color=_HEADER_TEXT, size=10)
        cell.alignment = Alignment(vertical="center", horizontal="left")
    sheet.row_dimensions[1].height = 22

    widths = [5, 12, 12, 14, 34, 22, 18, 18, 18, 18, 12, 24, 12]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for offset in range(len(outcome.property_columns)):
        sheet.column_dimensions[get_column_letter(len(BASE_HEADERS) + 1 + offset)].width = 18

    # Titik di luar KPS diberi warna supaya langsung kelihatan saat ditelusuri
    # ribuan baris -- ini justru informasi yang paling dicari pengguna.
    status_column = BASE_HEADERS.index("Status") + 1
    for row_number in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row_number, column=status_column)
        if cell.value == STATUS_OUTSIDE:
            cell.font = Font(bold=True, color=_OUTSIDE, size=10)

    sheet.freeze_panes = "A2"
    if outcome.points:
        last_column = get_column_letter(len(headers))
        sheet.auto_filter.ref = f"A1:{last_column}{len(outcome.points) + 1}"


def _write_section(
    sheet,
    start_row: int,
    title: str,
    label_header: str,
    data: list[dict[str, Any]],
    total: int,
    limit: int = 15,
) -> tuple[int, int, int]:
    """Tulis satu tabel ringkasan.

    Mengembalikan (baris_kosong_berikutnya, baris_header, jumlah_baris_data)
    supaya pemanggil bisa menempelkan grafik ke rentang yang tepat.
    """
    title_cell = sheet.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12, color=_INK)

    header_row = start_row + 1
    header_fill = PatternFill("solid", fgColor=_INK)
    for offset, header in enumerate((label_header, "Jumlah", "%")):
        cell = sheet.cell(row=header_row, column=1 + offset, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color=_HEADER_TEXT, size=10)
        cell.border = _CELL_BORDER

    shown = data[:limit]
    for index, item in enumerate(shown):
        row_number = header_row + 1 + index
        share = (item["count"] / total) if total else 0
        for offset, value in enumerate((item["label"], item["count"], share)):
            cell = sheet.cell(row=row_number, column=1 + offset, value=value)
            cell.border = _CELL_BORDER
            if index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=_BAND)
        sheet.cell(row=row_number, column=3).number_format = "0.0%"

    if len(data) > limit:
        note_row = header_row + 1 + len(shown)
        note = sheet.cell(
            row=note_row,
            column=1,
            value=f"... dan {len(data) - limit} lainnya (lihat sheet {DATA_SHEET_TITLE})",
        )
        note.font = Font(italic=True, size=9, color=_MUTED)

    next_row = header_row + max(len(shown), 1) + (2 if len(data) <= limit else 3)
    return next_row, header_row, len(shown)


def _attach_bar_chart(sheet, anchor: str, title: str, header_row: int, row_count: int) -> None:
    if row_count <= 0:
        return
    chart = BarChart()
    chart.type = "bar"
    chart.title = title
    chart.height = min(2 + row_count * 0.65, 11)
    chart.width = 16
    chart.legend = None
    chart.y_axis.majorGridlines = None
    data = Reference(sheet, min_col=2, min_row=header_row, max_row=header_row + row_count)
    categories = Reference(sheet, min_col=1, min_row=header_row + 1, max_row=header_row + row_count)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, anchor)


def _write_dashboard_sheet(sheet, outcome: MatchOutcome, source_name: str) -> None:
    summary = outcome.summary

    sheet["A1"] = "DASHBOARD PENCOCOKAN TITIK KE KAWASAN PERHUTANAN SOSIAL"
    sheet["A1"].font = Font(bold=True, size=16, color=_INK)
    sheet["A2"] = (
        f"Sumber berkas: {source_name} ({outcome.source_format})  •  "
        f"Dibuat {datetime.now(WIB).strftime('%d-%m-%Y %H:%M')} WIB"
    )
    sheet["A2"].font = Font(size=10, color=_MUTED, italic=True)

    cards = [
        ("Total Titik", summary.total_points),
        ("Masuk KPS", summary.inside_count),
        ("Di Luar KPS", summary.outside_count),
        ("KPS Terdampak", summary.distinct_kps),
    ]
    for offset, (label, value) in enumerate(cards):
        column = 1 + offset * 2
        label_cell = sheet.cell(row=4, column=column, value=label)
        label_cell.font = Font(bold=True, size=9, color=_MUTED)
        value_cell = sheet.cell(row=5, column=column, value=value)
        value_cell.font = Font(
            bold=True, size=20, color=_OUTSIDE if label == "Di Luar KPS" and value else _INK
        )
        sheet.cell(row=6, column=column).fill = PatternFill("solid", fgColor=_ACCENT)

    row = 8
    if outcome.warnings:
        for warning in outcome.warnings:
            cell = sheet.cell(row=row, column=1, value=f"Catatan: {warning}")
            cell.font = Font(size=9, color=_OUTSIDE, italic=True)
            row += 1
        row += 1

    total = summary.total_points
    row, header_row, count = _write_section(sheet, row, "Titik per KPS", "KPS", summary.by_kps, total)
    _attach_bar_chart(sheet, f"E{header_row}", "Titik per KPS", header_row, count)

    row, header_row, count = _write_section(
        sheet, row, "Titik per Balai PS", "Balai PS", summary.by_wilker, total
    )
    _attach_bar_chart(sheet, f"E{header_row}", "Titik per Balai PS", header_row, count)

    row, header_row, count = _write_section(
        sheet, row, "Titik per Provinsi", "Provinsi", summary.by_province, total
    )
    _attach_bar_chart(sheet, f"E{header_row}", "Titik per Provinsi", header_row, count)

    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 10
    sheet.column_dimensions["D"].width = 3


def build_excel_file(outcome: MatchOutcome, source_name: str) -> bytes:
    workbook = Workbook()

    data_sheet = workbook.active
    data_sheet.title = DATA_SHEET_TITLE
    _write_data_sheet(data_sheet, outcome)

    dashboard_sheet = workbook.create_sheet(DASHBOARD_SHEET_TITLE)
    _write_dashboard_sheet(dashboard_sheet, outcome, source_name)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------------- PDF


def _escape(value: Any) -> str:
    text = "" if value is None else str(value)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _summary_table_html(title: str, label_header: str, data: list[dict[str, Any]], total: int) -> str:
    if not data:
        return ""
    # Semua baris ditampilkan -- sebelumnya dipotong 15 baris + "... dan N
    # lainnya", tapi jumlah entitas (KPS/Balai PS/Provinsi) selalu jauh lebih
    # kecil dari jumlah titik hotspot, jadi tidak ada risiko tabel meledak
    # seperti kalau ini daftar titik mentah.
    rows = "".join(
        "<tr>"
        f"<td>{_escape(item['label'])}</td>"
        f"<td class='num'>{item['count']}</td>"
        f"<td class='num'>{(item['count'] / total * 100 if total else 0):.1f}%</td>"
        "</tr>"
        for item in data
    )
    return f"""
    <h2>{_escape(title)}</h2>
    <table>
      <thead><tr><th>{_escape(label_header)}</th><th class='num'>Jumlah</th><th class='num'>%</th></tr></thead>
      <tbody>{rows}</tbody>
    </table>
    """


def _iter_rings(geometry: dict[str, Any]) -> list[list[tuple[float, float]]]:
    """Ratakan Polygon/MultiPolygon jadi daftar cincin koordinat (lon, lat)."""
    kind = geometry.get("type")
    coordinates = geometry.get("coordinates")
    rings: list[list[tuple[float, float]]] = []

    def _add(ring: Any) -> None:
        if not isinstance(ring, (list, tuple)):
            return
        points = [
            (float(pair[0]), float(pair[1]))
            for pair in ring
            if isinstance(pair, (list, tuple)) and len(pair) >= 2
        ]
        if len(points) >= 3:
            rings.append(points)

    if kind == "Polygon" and isinstance(coordinates, (list, tuple)):
        for ring in coordinates:
            _add(ring)
    elif kind == "MultiPolygon" and isinstance(coordinates, (list, tuple)):
        for polygon in coordinates:
            if isinstance(polygon, (list, tuple)):
                for ring in polygon:
                    _add(ring)
    return rings


def _kps_map_svg(
    geometry: dict[str, Any],
    points: list[Any],
    *,
    width: int = 190,
    height: int = 130,
    padding: int = 8,
) -> str:
    """Gambar satu KPS: batas kawasan + titik hotspot di dalamnya.

    Dibuat sebagai SVG inline, bukan gambar raster, supaya tidak menambah
    dependensi dan tetap tajam berapa pun perbesarannya saat PDF dicetak.
    """
    rings = _iter_rings(geometry)
    if not rings:
        return ""

    xs = [x for ring in rings for x, _ in ring] + [p.longitude for p in points]
    ys = [y for ring in rings for _, y in ring] + [p.latitude for p in points]
    min_x, max_x = min(xs), max(xs)
    min_y, max_y = min(ys), max(ys)

    span_x = max(max_x - min_x, 1e-9)
    span_y = max(max_y - min_y, 1e-9)

    # Satu skala untuk kedua sumbu supaya bentuk kawasan tidak gepeng.
    # Lintang dikoreksi cos(lat) supaya proporsinya mendekati jarak sebenarnya.
    import math

    mid_lat = (min_y + max_y) / 2
    lon_scale = math.cos(math.radians(mid_lat)) or 1.0
    scale = min((width - 2 * padding) / (span_x * lon_scale), (height - 2 * padding) / span_y)

    draw_w = span_x * lon_scale * scale
    draw_h = span_y * scale
    offset_x = (width - draw_w) / 2
    offset_y = (height - draw_h) / 2

    def project(lon: float, lat: float) -> tuple[float, float]:
        x = offset_x + (lon - min_x) * lon_scale * scale
        # SVG menghitung y dari atas, lintang dari bawah -- karena itu dibalik.
        y = offset_y + (max_y - lat) * scale
        return x, y

    paths = []
    for ring in rings:
        coords = " ".join(f"{px:.2f},{py:.2f}" for px, py in (project(lon, lat) for lon, lat in ring))
        paths.append(f'<polygon points="{coords}" />')

    dots = []
    for point in points:
        px, py = project(point.longitude, point.latitude)
        dots.append(f'<circle cx="{px:.2f}" cy="{py:.2f}" r="2.6" />')

    return (
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" '
        f'viewBox="0 0 {width} {height}">'
        f'<rect width="{width}" height="{height}" fill="#F7F9F8" stroke="#D8DEDA"/>'
        f'<g fill="#DCE8DF" stroke="#1B3A2B" stroke-width="0.9" stroke-linejoin="round">'
        f'{"".join(paths)}</g>'
        f'<g fill="#E0862A" stroke="#8A4B12" stroke-width="0.6">{"".join(dots)}</g>'
        "</svg>"
    )


def _kps_maps_html(outcome: MatchOutcome) -> str:
    """Satu kartu peta per KPS, diurutkan dari yang titiknya terbanyak.

    Semua KPS terdampak ditampilkan -- sebelumnya dipotong 12 kartu teratas
    + catatan "menampilkan N dari M", tapi jumlah KPS terdampak per unggahan
    selalu jauh lebih kecil dari jumlah titik, jadi aman ditampilkan penuh.
    """
    if not outcome.polygon_geometries:
        return ""

    grouped: dict[int, list[Any]] = {}
    names: dict[int, str] = {}
    for point in outcome.points:
        if not point.kps:
            continue
        polygon_id = point.kps.get("polygon_metadata_id")
        if polygon_id is None:
            continue
        key = int(polygon_id)
        grouped.setdefault(key, []).append(point)
        names.setdefault(key, str(point.kps.get("lembaga") or "(tanpa nama)"))

    ranked = sorted(grouped.items(), key=lambda kv: (-len(kv[1]), names.get(kv[0], "")))
    cards = []
    for polygon_id, points in ranked:
        geometry = outcome.polygon_geometries.get(polygon_id)
        if not geometry:
            continue
        svg = _kps_map_svg(geometry, points)
        if not svg:
            continue
        wilker = _escape(points[0].kps.get("wilker_bps"))
        cards.append(
            "<div class='mapcard'>"
            f"{svg}"
            f"<div class='mapcard__name'>{_escape(names.get(polygon_id, ''))}</div>"
            f"<div class='mapcard__meta'>{wilker} &bull; {len(points)} hotspot</div>"
            "</div>"
        )

    if not cards:
        return ""

    # Bagian peta selalu dimulai di halaman baru. Kartu peta berukuran tetap dan
    # tidak bisa dipotong, jadi kalau dibiarkan mengalir ia hampir selalu tidak
    # muat di sisa halaman dan meninggalkan judulnya sendirian di bawah.
    return (
        "<div class='mapsection'>"
        "<h2>Peta Sebaran Hotspot per KPS</h2>"
        "<p class='more'>Area hijau = batas kawasan KPS; titik oranye = hotspot di dalamnya. "
        "Skala tiap peta menyesuaikan luas kawasannya masing-masing.</p>"
        f"<div class='mapgrid'>{''.join(cards)}</div>"
        "</div>"
    )


def build_pdf_file(outcome: MatchOutcome, source_name: str) -> bytes:
    from weasyprint import HTML

    summary = outcome.summary
    generated = datetime.now(WIB).strftime("%d-%m-%Y %H:%M")

    warnings_html = ""
    if outcome.warnings:
        items = "".join(f"<li>{_escape(w)}</li>" for w in outcome.warnings)
        warnings_html = f"<div class='warn'><strong>Catatan:</strong><ul>{items}</ul></div>"

    # Rincian hanya memuat hotspot yang masuk KPS -- itu yang dipakai untuk
    # tindak lanjut. Jumlah yang di luar KPS tetap dilaporkan di kartu ringkasan
    # supaya angkanya tidak hilang, hanya barisnya yang tidak dirinci di sini.
    # Karena semua baris kini berstatus sama, kolom Status ikut dihapus: kolom
    # yang isinya selalu sama hanya menambah lebar tanpa menambah informasi.
    inside_points = [point for point in outcome.points if point.inside_kps]

    # Tabel rinci dibatasi supaya PDF tetap wajar dibuka; data lengkap ada di Excel.
    detail_limit = 200
    detail_rows = "".join(
        "<tr>"
        f"<td class='num'>{index}</td>"
        f"<td class='num'>{point.latitude:.5f}</td>"
        f"<td class='num'>{point.longitude:.5f}</td>"
        f"<td>{_escape(kategori_confidence)}</td>"
        f"<td>{_escape(kategori_frp)}</td>"
        f"<td>{_escape((point.kps or {}).get('lembaga'))}</td>"
        f"<td>{_escape((point.kps or {}).get('wilker_bps'))}</td>"
        f"<td>{_escape((point.kps or {}).get('nama_prov'))}</td>"
        "</tr>"
        for index, (point, kategori_confidence, kategori_frp) in enumerate(
            (
                (point, *_confidence_and_frp_category(point.properties))
                for point in inside_points[:detail_limit]
            ),
            start=1,
        )
    )
    detail_note = (
        f"<p class='more'>Menampilkan {detail_limit} dari {len(inside_points):,} hotspot di KPS. "
        "Data lengkap (termasuk hotspot di luar KPS) tersedia di berkas Excel.</p>"
        if len(inside_points) > detail_limit
        else ""
    )
    if not inside_points:
        detail_rows = (
            "<tr><td colspan='8' class='empty'>"
            "Tidak ada hotspot yang masuk kawasan KPS."
            "</td></tr>"
        )

    maps_html = _kps_maps_html(outcome)

    html = f"""<!doctype html>
<html lang="id"><head><meta charset="utf-8"><title>Laporan Hotspot pada Persetujuan Perhutanan Sosial</title>
<style>
  @page {{ size: A4; margin: 16mm 14mm; }}
  body {{ font-family: "DejaVu Sans", sans-serif; color: #1B3A2B; font-size: 9pt; }}
  h1 {{ font-size: 16pt; margin: 0 0 2mm; }}
  /* Judul bagian ikut pindah bersama isinya; tanpa ini judul bisa tertinggal
     sendirian di ujung halaman sementara tabel/petanya lompat ke halaman
     berikutnya, menyisakan ruang kosong besar. */
  h2 {{ font-size: 11pt; margin: 6mm 0 2mm; border-bottom: 1px solid #D8DEDA;
        padding-bottom: 1mm; break-after: avoid; page-break-after: avoid; }}
  .sub {{ color: #6B726D; font-style: italic; margin: 0 0 5mm; }}
  .cards {{ display: flex; gap: 4mm; margin-bottom: 5mm; }}
  .card {{ flex: 1; border: 1px solid #D8DEDA; border-top: 3px solid #E0862A; padding: 3mm; }}
  .card .label {{ font-size: 7.5pt; color: #6B726D; text-transform: uppercase; }}
  .card .value {{ font-size: 16pt; font-weight: bold; }}
  .card.alert .value {{ color: #B4453C; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th {{ background: #1B3A2B; color: #fff; font-size: 8pt; text-align: left; padding: 1.6mm 2mm; }}
  td {{ border: 1px solid #D8DEDA; padding: 1.4mm 2mm; font-size: 8pt; }}
  tbody tr:nth-child(even) td {{ background: #EEF3EF; }}
  .num {{ text-align: right; }}
  .outside {{ color: #B4453C; font-weight: bold; }}
  .inside {{ color: #1B3A2B; }}
  .more {{ font-size: 8pt; color: #6B726D; font-style: italic; }}
  .empty {{ text-align: center; color: #6B726D; font-style: italic; padding: 4mm 2mm; }}
  /* Sengaja inline-block, bukan flex: WeasyPrint memperlakukan kontainer flex
     sebagai satu blok utuh yang tidak bisa dipecah, sehingga seluruh grid
     melompat ke halaman berikutnya dan meninggalkan judulnya sendirian. */
  .mapsection {{ break-before: page; page-break-before: always; }}
  .mapgrid {{ margin-top: 2mm; font-size: 0; }}
  .mapcard {{ display: inline-block; vertical-align: top; width: 52mm;
              border: 1px solid #D8DEDA; padding: 2mm; margin: 0 2.5mm 2.5mm 0;
              font-size: 8pt; break-inside: avoid; page-break-inside: avoid; }}
  .mapcard svg {{ display: block; width: 100%; height: auto; }}
  .mapcard__name {{ font-size: 7.5pt; font-weight: bold; margin-top: 1.5mm;
                    word-wrap: break-word; }}
  .mapcard__meta {{ font-size: 7pt; color: #6B726D; }}
  .warn {{ border-left: 3px solid #B4453C; background: #FBF0EF; padding: 2mm 3mm; margin-bottom: 4mm; font-size: 8.5pt; }}
  .warn ul {{ margin: 1mm 0 0 4mm; padding: 0; }}
</style></head>
<body>
  <h1>Laporan Hotspot pada Persetujuan Perhutanan Sosial</h1>
  <p class="sub">Sumber berkas: {_escape(source_name)} ({_escape(outcome.source_format)}) &bull; Dibuat {generated} WIB</p>
  {warnings_html}
  <div class="cards">
    <div class="card"><div class="label">Total Hotspot</div><div class="value">{summary.total_points:,}</div></div>
    <div class="card"><div class="label">Hotspot di KPS</div><div class="value">{summary.inside_count:,}</div></div>
    <div class="card {'alert' if summary.outside_count else ''}"><div class="label">Hotspot di luar KPS</div><div class="value">{summary.outside_count:,}</div></div>
    <div class="card"><div class="label">KPS Terdampak</div><div class="value">{summary.distinct_kps:,}</div></div>
  </div>
  {_summary_table_html("Hotspot per KPS", "KPS", summary.by_kps, summary.total_points)}
  {_summary_table_html("Hotspot per Balai PS", "Balai PS", summary.by_wilker, summary.total_points)}
  {_summary_table_html("Hotspot per Provinsi", "Provinsi", summary.by_province, summary.total_points)}
  {maps_html}
  <h2>Rincian Hotspot di KPS</h2>
  <table>
    <thead><tr><th class="num">No</th><th class="num">Latitude</th><th class="num">Longitude</th>
    <th>Confidence</th><th>FRP</th><th>KPS</th><th>Balai PS</th><th>Provinsi</th></tr></thead>
    <tbody>{detail_rows}</tbody>
  </table>
  {detail_note}
</body></html>"""

    return HTML(string=html).write_pdf()
