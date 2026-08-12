from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.services.point_match_service import MatchedPoint, MatchOutcome, MatchSummary
from app.services.point_report_service import (
    BASE_HEADERS,
    DASHBOARD_SHEET_TITLE,
    DATA_SHEET_TITLE,
    STATUS_INSIDE,
    STATUS_OUTSIDE,
    build_excel_file,
    build_report_rows,
)


def _outcome(points, property_columns=None, warnings=None):
    inside = sum(1 for p in points if p.kps)
    return MatchOutcome(
        points=points,
        summary=MatchSummary(
            total_points=len(points),
            inside_count=inside,
            outside_count=len(points) - inside,
            distinct_kps=len({p.kps["lembaga"] for p in points if p.kps}),
            by_kps=[{"label": "LPHD DEMO", "count": inside}] if inside else [],
            by_wilker=[{"label": "Balai PS X", "count": inside}] if inside else [],
            by_province=[{"label": "Kalbar", "count": inside}] if inside else [],
        ),
        source_format="geojson",
        warnings=warnings or [],
        skipped_features=0,
        property_columns=property_columns or [],
    )


def _inside_point(**properties):
    return MatchedPoint(
        latitude=-1.88,
        longitude=110.26,
        properties=properties,
        kps={
            "lembaga": "LPHD DEMO",
            "wilker_bps": "Balai PS X",
            "nama_prov": "Kalbar",
            "nama_kab": "Ketapang",
            "nama_kec": "Kec A",
            "nama_desa": "Desa B",
            "skema": "HD",
            "no_sk": "SK.123",
            "tgl_sk": "2020-01-01",
        },
    )


def _outside_point(**properties):
    return MatchedPoint(latitude=0.0, longitude=90.0, properties=properties, kps=None)


def test_rows_mark_inside_and_outside_status():
    outcome = _outcome([_inside_point(), _outside_point()])
    rows = build_report_rows(outcome)

    status_index = BASE_HEADERS.index("Status")
    assert rows[0][status_index] == STATUS_INSIDE
    assert rows[1][status_index] == STATUS_OUTSIDE


def test_rows_append_user_metadata_after_fixed_columns():
    outcome = _outcome(
        [_inside_point(kode="A-1", regu="R2")],
        property_columns=["kode", "regu"],
    )
    rows = build_report_rows(outcome)

    assert len(rows[0]) == len(BASE_HEADERS) + 2
    assert rows[0][len(BASE_HEADERS)] == "A-1"
    assert rows[0][len(BASE_HEADERS) + 1] == "R2"


def test_outside_point_leaves_kps_columns_blank_not_missing():
    """Kolom KPS harus tetap ada (kosong) supaya jumlah kolom tiap baris sama --
    kalau tidak, Excel/CSV jadi bergeser."""
    outcome = _outcome([_outside_point(kode="B")], property_columns=["kode"])
    rows = build_report_rows(outcome)

    assert len(rows[0]) == len(BASE_HEADERS) + 1
    lembaga_index = BASE_HEADERS.index("KPS (Lembaga)")
    assert rows[0][lembaga_index] == ""


def test_excel_has_both_data_and_dashboard_sheets():
    outcome = _outcome([_inside_point(kode="A"), _outside_point(kode="B")], ["kode"])
    workbook = load_workbook(BytesIO(build_excel_file(outcome, "titik.geojson")))

    assert workbook.sheetnames == [DATA_SHEET_TITLE, DASHBOARD_SHEET_TITLE]


def test_excel_data_sheet_headers_include_user_columns():
    outcome = _outcome([_inside_point(kode="A")], ["kode"])
    workbook = load_workbook(BytesIO(build_excel_file(outcome, "titik.geojson")))
    sheet = workbook[DATA_SHEET_TITLE]

    headers = [cell.value for cell in sheet[1]]
    assert headers == [*BASE_HEADERS, "kode"]
    assert sheet.max_row == 2


def test_dashboard_sheet_reports_outside_count_and_has_charts():
    outcome = _outcome([_inside_point(), _outside_point(), _outside_point()])
    workbook = load_workbook(BytesIO(build_excel_file(outcome, "titik.geojson")))
    sheet = workbook[DASHBOARD_SHEET_TITLE]

    values = [cell.value for row in sheet.iter_rows() for cell in row]
    assert "Di Luar KPS" in values
    assert 2 in values  # dua titik di luar kawasan
    # Dashboard tanpa grafik bukan dashboard.
    assert len(sheet._charts) >= 1


def test_dashboard_surfaces_parser_warnings():
    outcome = _outcome([_inside_point()], warnings=["Berkas .prj tidak ada."])
    workbook = load_workbook(BytesIO(build_excel_file(outcome, "titik.zip")))
    sheet = workbook[DASHBOARD_SHEET_TITLE]

    text = " ".join(str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value)
    assert ".prj" in text


def test_pdf_is_generated_and_looks_like_a_pdf():
    pytest.importorskip("weasyprint")
    from app.services.point_report_service import build_pdf_file

    outcome = _outcome([_inside_point(kode="A"), _outside_point(kode="B")], ["kode"])
    pdf = build_pdf_file(outcome, "titik.geojson")

    assert pdf.startswith(b"%PDF")
    assert len(pdf) > 1000
