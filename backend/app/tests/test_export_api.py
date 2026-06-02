import asyncio


async def _request_export(query_string: bytes) -> tuple[int, list[dict]]:
    from app.main import create_app

    app = create_app()
    messages: list[dict] = []
    request_sent = False

    scope = {
        "type": "http",
        "asgi": {"version": "3.0"},
        "http_version": "1.1",
        "method": "GET",
        "scheme": "http",
        "path": "/api/export.xlsx",
        "raw_path": b"/api/export.xlsx",
        "query_string": query_string,
        "headers": [],
        "client": ("127.0.0.1", 12345),
        "server": ("testserver", 80),
    }

    async def receive() -> dict:
        nonlocal request_sent
        if request_sent:
            return {"type": "http.disconnect"}
        request_sent = True
        return {"type": "http.request", "body": b"", "more_body": False}

    async def send(message: dict) -> None:
        messages.append(message)

    await app(scope, receive, send)
    start = next(message for message in messages if message["type"] == "http.response.start")
    return start["status"], messages


async def _request_cache_refresh() -> tuple[int, bytes]:
    from app.main import create_app

    app = create_app()
    messages: list[dict] = []
    request_sent = False

    scope = {
        "type": "http",
        "asgi": {"version": "3.0"},
        "http_version": "1.1",
        "method": "POST",
        "scheme": "http",
        "path": "/api/cache/refresh",
        "raw_path": b"/api/cache/refresh",
        "query_string": b"",
        "headers": [],
        "client": ("127.0.0.1", 12345),
        "server": ("testserver", 80),
    }

    async def receive() -> dict:
        nonlocal request_sent
        if request_sent:
            return {"type": "http.disconnect"}
        request_sent = True
        return {"type": "http.request", "body": b"", "more_body": False}

    async def send(message: dict) -> None:
        messages.append(message)

    await app(scope, receive, send)
    start = next(message for message in messages if message["type"] == "http.response.start")
    body = next(message for message in messages if message["type"] == "http.response.body")
    return start["status"], body["body"]


def test_export_endpoint_returns_xlsx_attachment(monkeypatch) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    from app.core.config import get_settings

    get_settings.cache_clear()
    monkeypatch.setenv("SHP_DIR", "app/tests/fixtures/shp")
    monkeypatch.setenv("NASA_FIRMS_API_KEY", "")
    try:
        status, messages = asyncio.run(
            _request_export(
                b"start_at=2026-01-01T00:00:00Z&end_at=2026-05-26T23:59:59Z"
                b"&satellites=MODIS&active_layers=sample_area"
            )
        )
        start = next(message for message in messages if message["type"] == "http.response.start")
        body = next(message for message in messages if message["type"] == "http.response.body")

        assert status == 200
        assert (
            start["headers"]
            and (
                b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                in dict(start["headers"])[b"content-type"]
            )
        )
        assert b"attachment; filename=eta-seuneu-hotspots.xlsx" == dict(start["headers"])[
            b"content-disposition"
        ]

        workbook = load_workbook(BytesIO(body["body"]))
        sheet = workbook.active
        assert sheet["A1"].value == "No"
    finally:
        get_settings.cache_clear()


def test_export_endpoint_accepts_datetime_window(monkeypatch) -> None:
    from app.core.config import get_settings

    get_settings.cache_clear()
    monkeypatch.setenv("SHP_DIR", "app/tests/fixtures/shp")
    monkeypatch.setenv("NASA_FIRMS_API_KEY", "")

    try:
        status, messages = asyncio.run(
            _request_export(
                b"start_date=2026-01-01&end_date=2026-05-26"
                b"&satellites=MODIS&active_layers=sample_area"
            )
        )

        assert status == 200
        assert any(message["type"] == "http.response.body" for message in messages)
    finally:
        get_settings.cache_clear()


def test_cache_refresh_endpoint_returns_queued_status() -> None:
    import json

    status, body = asyncio.run(_request_cache_refresh())

    assert status == 200
    assert json.loads(body.decode("utf-8")) == {"status": "queued"}


async def _request_export_pdf(query_string: bytes) -> tuple[int, list[dict]]:
    from app.main import create_app

    app = create_app()
    messages: list[dict] = []
    request_sent = False

    scope = {
        "type": "http",
        "asgi": {"version": "3.0"},
        "http_version": "1.1",
        "method": "GET",
        "scheme": "http",
        "path": "/api/export.pdf",
        "raw_path": b"/api/export.pdf",
        "query_string": query_string,
        "headers": [],
        "client": ("127.0.0.1", 12345),
        "server": ("testserver", 80),
    }

    async def receive() -> dict:
        nonlocal request_sent
        if request_sent:
            return {"type": "http.disconnect"}
        request_sent = True
        return {"type": "http.request", "body": b"", "more_body": False}

    async def send(message: dict) -> None:
        messages.append(message)

    await app(scope, receive, send)
    start = next(message for message in messages if message["type"] == "http.response.start")
    return start["status"], messages


def test_export_endpoint_returns_pdf_attachment(monkeypatch) -> None:
    from app.core.config import get_settings

    get_settings.cache_clear()
    monkeypatch.setenv("SHP_DIR", "app/tests/fixtures/shp")
    monkeypatch.setenv("NASA_FIRMS_API_KEY", "")
    try:
        status, messages = asyncio.run(
            _request_export_pdf(
                b"start_at=2026-01-01T00:00:00Z&end_at=2026-05-26T23:59:59Z"
                b"&satellites=MODIS&active_layers=sample_area"
            )
        )
        start = next(message for message in messages if message["type"] == "http.response.start")
        body = next(message for message in messages if message["type"] == "http.response.body")

        assert status == 200
        assert (
            start["headers"]
            and (
                b"application/pdf"
                in dict(start["headers"])[b"content-type"]
            )
        )
        assert b"attachment; filename=eta-seuneu-hotspots-report.pdf" == dict(start["headers"])[
            b"content-disposition"
        ]
        assert len(body["body"]) > 0
    finally:
        get_settings.cache_clear()


def test_export_endpoints_apply_matrix_filters(monkeypatch) -> None:
    from io import BytesIO
    from openpyxl import load_workbook
    from app.services.hotspot_service import HotspotService
    from app.core.config import get_settings

    get_settings.cache_clear()
    monkeypatch.setenv("SHP_DIR", "app/tests/fixtures/shp")
    monkeypatch.setenv("NASA_FIRMS_API_KEY", "")

    mock_hotspots = [
        {
            "latitude": -8.4071,
            "longitude": 118.75785,
            "brightness": 299.28,
            "frp": 0.26,
            "confidence": "n",
            "source": "VIIRS NOAA-20",
            "satellite": "N20",
            "daynight": "N",
            "layer_id": "PS_FEB_26",
            "layer_name": "KTH KAPENTA NANGANAE",
            "agency_name": "KTH KAPENTA NANGANAE",
            "detected_at": "2026-05-28T17:24:00Z",
            "province_name": "Nusa Tenggara Barat",
            "polygon_metadata": {
                "WILKER_BPS": "Balai PS Denpasar",
                "LEMBAGA": "KTH KAPENTA NANGANAE",
            }
        },
        {
            "latitude": 0.81682,
            "longitude": 110.12852,
            "brightness": 327.5,
            "frp": 35.0,
            "confidence": "80",
            "source": "MODIS",
            "satellite": "Terra",
            "daynight": "D",
            "layer_id": "PS_FEB_26",
            "layer_name": "LDPH ANDU LESTARI",
            "agency_name": "LDPH ANDU LESTARI",
            "detected_at": "2026-05-29T05:44:00Z",
            "province_name": "Kalimantan Barat",
            "polygon_metadata": {
                "WILKER_BPS": "Balai PS Banjarbaru",
                "LEMBAGA": "LDPH ANDU LESTARI",
            }
        }
    ]

    async def mock_fetch_filtered_hotspots(self, query, bypass_cache: bool = False) -> dict:
        return {
            "count": len(mock_hotspots),
            "hotspots": mock_hotspots,
            "stats": {}
        }

    monkeypatch.setattr(HotspotService, "fetch_filtered_hotspots", mock_fetch_filtered_hotspots)

    try:
        # 1. Test XLSX export with province filter
        status, messages = asyncio.run(
            _request_export(
                b"start_date=2026-01-01&end_date=2026-05-26"
                b"&satellites=MODIS&active_layers=sample_area"
                b"&province=Nusa Tenggara Barat"
            )
        )
        assert status == 200
        body = next(message for message in messages if message["type"] == "http.response.body")
        wb = load_workbook(BytesIO(body["body"]))
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][1] == "KTH KAPENTA NANGANAE"

        # 2. Test XLSX export with wilker filter
        status, messages = asyncio.run(
            _request_export(
                b"start_date=2026-01-01&end_date=2026-05-26"
                b"&satellites=MODIS&active_layers=sample_area"
                b"&wilker=Balai PS Banjarbaru"
            )
        )
        assert status == 200
        body = next(message for message in messages if message["type"] == "http.response.body")
        wb = load_workbook(BytesIO(body["body"]))
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][1] == "LDPH ANDU LESTARI"

        # 3. Test XLSX export with confidence filter
        status, messages = asyncio.run(
            _request_export(
                b"start_date=2026-01-01&end_date=2026-05-26"
                b"&satellites=MODIS&active_layers=sample_area"
                b"&confidence=Rendah"
            )
        )
        assert status == 200
        body = next(message for message in messages if message["type"] == "http.response.body")
        wb = load_workbook(BytesIO(body["body"]))
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][1] == "KTH KAPENTA NANGANAE"

        # 4. Test PDF export with filters
        status, messages = asyncio.run(
            _request_export_pdf(
                b"start_date=2026-01-01&end_date=2026-05-26"
                b"&satellites=MODIS&active_layers=sample_area"
                b"&confidence=Rendah"
            )
        )
        assert status == 200
        start = next(message for message in messages if message["type"] == "http.response.start")
        body = next(message for message in messages if message["type"] == "http.response.body")
        assert b"application/pdf" in dict(start["headers"])[b"content-type"]
        assert len(body["body"]) > 0
    finally:
        get_settings.cache_clear()


