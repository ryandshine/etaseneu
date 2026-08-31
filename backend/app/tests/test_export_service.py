def test_build_export_rows_maps_hotspot_fields() -> None:
    from app.services.export_service import build_export_rows

    rows = build_export_rows(
        [
            {
                "layer_name": "sample_area",
                "source": "MODIS",
                "detected_at": "2026-05-24T06:12:00",
                "latitude": 4.1,
                "longitude": 95.1,
                "confidence": "Rendah",
                "brightness": 330.4,
            }
        ]
    )

    assert rows[0]["Nama Wilayah"] == "sample_area"
    assert rows[0]["Satelit"] == "MODIS"


def test_build_excel_file_writes_headers_and_row_values() -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    from app.services.export_service import build_excel_file

    content = build_excel_file(
        [
            {
                "layer_name": "sample_area",
                "source": "MODIS",
                "detected_at": "2026-05-24T06:12:00",
                "latitude": 4.1,
                "longitude": 95.1,
                "confidence": "Rendah",
                "brightness": 330.4,
            }
        ]
    )

    workbook = load_workbook(BytesIO(content))
    sheet = workbook["Data Hotspot"]

    assert [cell.value for cell in sheet[1]] == [
        "No",
        "Nama Wilayah",
        "No. SK",
        "Tgl SK",
        "Skema",
        "Balai PS",
        "Provinsi",
        "Kabupaten",
        "Satelit",
        "Tanggal Deteksi",
        "Latitude",
        "Longitude",
        "Confidence",
        "Kategori FRP",
        "FRP (MW)",
        "Brightness",
        "Fungsi Kawasan Hutan",
        "Nama Kawasan",
        "Kelompok",
    ]
    # Sel kosong ("") dibaca kembali oleh openpyxl sebagai None.
    # detected_at tanpa offset diperlakukan sebagai UTC (sama seperti yang
    # ditulis hotspot_normalizer), jadi 06:12 UTC -> 13:12 WIB. Sebelumnya
    # datetime naive ikut zona waktu server, sehingga hasilnya berbeda antara
    # mesin dev (WIB) dan container produksi (UTC).
    assert [cell.value for cell in sheet[2]] == [
        1,
        "sample_area",
        None,
        None,
        "Tanpa Skema",
        "Tanpa Balai",
        "Tanpa Provinsi",
        None,
        "MODIS",
        "24-05-2026 13:12 WIB",
        4.1,
        95.1,
        "Rendah",
        "Rendah",
        None,
        330.4,
        None,
        None,
        None,
    ]


def _sample_hotspots() -> list[dict]:
    return [
        {
            "layer_name": "LPHD KALIBANDUNG",
            "source": "VIIRS NOAA-21",
            "detected_at": "2026-07-24T06:12:00Z",
            "latitude": -0.5,
            "longitude": 109.4,
            "confidence": "h",
            "frp": 45.0,
            "brightness": 340.0,
            "province_name": "Kalimantan Barat",
            "polygon_metadata": {
                "WILKER_BPS": "Balai PS Banjarbaru",
                "NAMA_PROV": "Kalimantan Barat",
                "NAMA_KAB": "Kubu Raya",
                "SKEMA": "PPHD",
            },
        },
        {
            "layer_name": "LK REP",
            "source": "MODIS",
            "detected_at": "2026-08-02T06:12:00Z",
            "latitude": -7.1,
            "longitude": 140.2,
            "confidence": "20",
            "frp": 4.0,
            "brightness": 310.0,
            "province_name": "Papua Selatan",
            "polygon_metadata": {
                "WILKER_BPS": "Balai PS Manokwari",
                "NAMA_PROV": "Papua Selatan",
                "NAMA_KAB": "Merauke",
                "SKEMA": "PKK",
            },
        },
    ]


def test_export_rows_include_balai_and_provinsi() -> None:
    from app.services.export_service import build_export_rows

    rows = build_export_rows(_sample_hotspots())

    assert rows[0]["Balai PS"] == "Balai PS Banjarbaru"
    assert rows[0]["Provinsi"] == "Kalimantan Barat"
    assert rows[0]["Kabupaten"] == "Kubu Raya"
    assert rows[1]["Balai PS"] == "Balai PS Manokwari"
    assert rows[1]["Provinsi"] == "Papua Selatan"


def test_export_rows_include_fungsi_kawasan_hutan() -> None:
    from app.services.export_service import build_export_rows

    rows = build_export_rows(
        [
            {
                "layer_name": "LPHD X",
                "source": "MODIS",
                "detected_at": "2026-08-02T06:12:00Z",
                "latitude": -1.0,
                "longitude": 101.0,
                "kawasan_hutan": {
                    "kode": 100300,
                    "fungsi": "Hutan Lindung",
                    "singkatan": "HL",
                    "nama_kawasan": "Air Bangis",
                    "kelompok": "Lindung",
                },
            },
            {"layer_name": "LPHD Y", "source": "MODIS", "detected_at": "2026-08-02T06:12:00Z"},
        ]
    )

    assert rows[0]["Fungsi Kawasan Hutan"] == "Hutan Lindung"
    assert rows[0]["Nama Kawasan"] == "Air Bangis"
    assert rows[0]["Kelompok"] == "Lindung"
    # Titik tanpa atribusi kawasan -> kolom kosong, bukan error.
    assert rows[1]["Fungsi Kawasan Hutan"] == ""
    assert rows[1]["Kelompok"] == ""


def test_provinsi_falls_back_to_polygon_metadata_when_province_name_missing() -> None:
    from app.services.export_service import build_export_rows

    rows = build_export_rows(
        [{"layer_name": "X", "polygon_metadata": {"NAMA_PROV": "Riau"}}]
    )

    assert rows[0]["Provinsi"] == "Riau"


def test_dashboard_summary_aggregates_by_balai_provinsi_and_month() -> None:
    from app.services.export_service import build_dashboard_summary

    summary = build_dashboard_summary(_sample_hotspots())

    assert summary["total"] == 2
    assert summary["jumlah_balai"] == 2
    assert summary["jumlah_provinsi"] == 2
    assert dict(summary["per_balai"]) == {
        "Balai PS Banjarbaru": 1,
        "Balai PS Manokwari": 1,
    }
    assert dict(summary["per_provinsi"]) == {
        "Kalimantan Barat": 1,
        "Papua Selatan": 1,
    }
    # detected_at UTC dikonversi ke WIB dulu sebelum dikelompokkan per bulan.
    assert summary["per_bulan"] == [("2026-07", 1), ("2026-08", 1)]
    assert dict(summary["per_confidence"])["Tinggi"] == 1
    assert dict(summary["per_frp"])["Tinggi"] == 1


def test_build_excel_file_creates_dashboard_sheet() -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    from app.services.export_service import build_excel_file

    workbook = load_workbook(BytesIO(build_excel_file(_sample_hotspots())))

    assert workbook.sheetnames == ["Dashboard", "Skema per Provinsi", "Data Hotspot"]
    dashboard = workbook["Dashboard"]
    assert "DASHBOARD" in str(dashboard["A1"].value)
    # Kartu ringkasan: label di baris 5, angkanya di baris 6.
    assert dashboard["A5"].value == "TOTAL HOTSPOT"
    assert dashboard["A6"].value == 2
    assert dashboard["G6"].value == 2  # Balai PS (kolom G:I)


def test_build_excel_file_handles_empty_hotspots() -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    from app.services.export_service import build_excel_file

    workbook = load_workbook(BytesIO(build_excel_file([])))

    assert workbook.sheetnames == ["Dashboard", "Skema per Provinsi", "Data Hotspot"]
    assert workbook["Dashboard"]["A6"].value == 0


def test_sk_number_collapses_duplicate_lines_from_source_shapefile() -> None:
    from app.services.polygon_fields import sk_number

    # 14 polygon di sumber menuliskan nomor yang sama dua kali dipisah CRLF.
    duplicated = {"polygon_metadata": {"NO_SK": "607/EKBANG/2016\r\n607/EKBANG/2016"}}
    assert sk_number(duplicated) == "607/EKBANG/2016"


def test_sk_number_keeps_genuinely_different_decree_numbers() -> None:
    from app.services.polygon_fields import sk_number

    multi = {"polygon_metadata": {"NO_SK": "343 Tahun 2010\n351 Tahun 2010"}}
    assert sk_number(multi) == "343 Tahun 2010 351 Tahun 2010"


def test_sk_number_is_blank_when_source_has_no_decree() -> None:
    from app.services.polygon_fields import sk_number

    assert sk_number({"polygon_metadata": {}}) == ""
    assert sk_number({}) == ""


def test_export_rows_include_sk_columns() -> None:
    from app.services.export_service import build_export_rows

    rows = build_export_rows(_sample_hotspots())

    assert "No. SK" in rows[0]
    assert "Tgl SK" in rows[0]


def test_export_rows_include_skema_column() -> None:
    from app.services.export_service import build_export_rows

    rows = build_export_rows(_sample_hotspots())

    assert rows[0]["Skema"] == "PPHD"
    assert rows[1]["Skema"] == "PKK"


def test_export_rows_label_missing_skema_instead_of_dropping_the_point() -> None:
    from app.services.export_service import build_export_rows

    rows = build_export_rows([{"layer_name": "X", "polygon_metadata": {"NAMA_PROV": "Riau"}}])

    assert rows[0]["Skema"] == "Tanpa Skema"


def test_dashboard_summary_aggregates_per_skema_and_skema_per_provinsi() -> None:
    from app.services.export_service import build_dashboard_summary

    summary = build_dashboard_summary(_sample_hotspots())

    assert summary["jumlah_skema"] == 2
    assert dict(summary["per_skema"]) == {"PPHD": 1, "PKK": 1}

    matrix = summary["skema_per_provinsi"]
    assert set(matrix.skema) == {"PPHD", "PKK"}
    assert {row.provinsi: row.total for row in matrix.rows} == {
        "Kalimantan Barat": 1,
        "Papua Selatan": 1,
    }
    assert matrix.grand_total == 2


def test_build_excel_file_writes_skema_per_provinsi_crosstab() -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    from app.services.export_service import build_excel_file

    workbook = load_workbook(BytesIO(build_excel_file(_sample_hotspots())))
    sheet = workbook["Skema per Provinsi"]

    # Baris 4 = header tabel silang: Provinsi, kolom per skema, lalu Total.
    header = [cell.value for cell in sheet[4] if cell.value is not None]
    assert header[0] == "Provinsi"
    assert header[-1] == "Total"
    assert set(header[1:-1]) == {"PPHD", "PKK"}

    body = {
        row[0]: row[1:]
        for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, values_only=True)
    }
    assert body["Kalimantan Barat"][-1] == 1
    assert body["Papua Selatan"][-1] == 1
    # Baris terakhir adalah total keseluruhan.
    assert body["Total"][-1] == 2


def test_dashboard_summary_includes_top_kps_detail() -> None:
    from app.services.export_service import build_dashboard_summary

    summary = build_dashboard_summary(_sample_hotspots())
    detail = summary["top_kps_detail"]

    assert len(detail) == 2
    assert detail[0]["rank"] == 1
    assert detail[0]["name"] == "LPHD KALIBANDUNG"
    assert detail[0]["provinsi"] == "Kalimantan Barat"
    assert detail[0]["skema"] == "PPHD"
    assert detail[0]["balai"] == "Balai PS Banjarbaru"
    assert detail[0]["count"] == 1
    assert detail[0]["percent"] == 0.5


def test_dashboard_sheet_contains_kpi_charts_and_tables() -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    from app.services.export_service import build_excel_file

    workbook = load_workbook(BytesIO(build_excel_file(_sample_hotspots())))
    dashboard = workbook["Dashboard"]

    # 1. Judul & Banner
    assert "ETA SENEU" in str(dashboard["A1"].value)
    assert "Periode Data" in str(dashboard["A3"].value)

    # 2. KPI Cards
    assert dashboard["A5"].value == "TOTAL HOTSPOT"
    assert dashboard["A6"].value == 2
    assert dashboard["D5"].value == "KPS TERDAMPAK"
    assert dashboard["D6"].value == 2
    assert dashboard["G5"].value == "BALAI PS"
    assert dashboard["G6"].value == 2

    # 3. Section Headers
    assert "1. SEBARAN HOTSPOT PER BALAI PS" in str(dashboard["A10"].value)
    assert "2. SEBARAN HOTSPOT PER SKEMA PERHUTANAN SOSIAL" in str(dashboard["I10"].value)
    assert "3. TREN DETEKSI HOTSPOT BULANAN" in str(dashboard["A29"].value)
    assert "4. DISTRIBUSI SEBARAN PER PROVINSI" in str(dashboard["I29"].value)
    assert "5. PARAMETER TINGKAT RISIKO" in str(dashboard["A48"].value)
    assert "6. 10 UNIT KPS PRIORITAS PENANGANAN" in str(dashboard["I48"].value)

    # 4. Tables in Section 3
    assert dashboard["A50"].value == "Tingkat Kepercayaan"
    assert dashboard["A56"].value == "Kategori FRP"
    assert dashboard["E50"].value == "Satelit Sensor"
    assert dashboard["I50"].value == "No"
    assert dashboard["J50"].value == "Nama Kawasan / KPS"
    assert dashboard["I51"].value == "#1"
    assert dashboard["J51"].value == "LPHD KALIBANDUNG"

    # 5. Charts attached
    assert len(dashboard._charts) == 4



_BURNED_REPORT = {
    "rows": [
        {"lembaga": "KOPERASI A", "skema": "PPHKm", "nama_prov": "Riau", "wilker_bps": "BPS Kampar",
         "burned_area_ha": 1786.3, "burned_months": 2, "latest_period": "2026-04", "is_estimated": False},
        {"lembaga": "KT C", "skema": "PPHKm", "nama_prov": "Lampung", "wilker_bps": "BPS Lampung",
         "burned_area_ha": 2.1, "burned_months": 1, "latest_period": "2026-03", "is_estimated": True},
    ],
    "by_skema": [{"skema": "PPHKm", "total_ha": 1788.4, "kps_count": 2}],
    "total_ha": 1788.4,
    "kps_count": 2,
}


def _load_sheet(content: bytes, title: str):
    import io

    import openpyxl

    return openpyxl.load_workbook(io.BytesIO(content))[title]


def test_build_excel_file_omits_burned_area_sheet_when_not_provided() -> None:
    """Pemanggil lain (mis. laporan Cek Titik) tidak mengirim data ini --
    sheet-nya tidak boleh muncul kosong di file mereka."""
    import io

    import openpyxl

    from app.services.export_service import build_excel_file

    workbook = openpyxl.load_workbook(io.BytesIO(build_excel_file([])))

    assert "Luas Terbakar" not in workbook.sheetnames


def test_build_excel_file_writes_burned_area_sheet() -> None:
    from app.services.export_service import BURNED_AREA_SHEET_TITLE, build_excel_file

    sheet = _load_sheet(
        build_excel_file([], burned_area_report=_BURNED_REPORT), BURNED_AREA_SHEET_TITLE
    )
    values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]

    assert any("REKAP LUAS BEKAS TERBAKAR" in str(value) for value in values)
    assert "KOPERASI A" in values
    assert 1786.3 in values
    # Baris yang lokasinya cuma perkiraan harus ditandai, bukan disamakan
    # dengan yang benar-benar terpetakan.
    assert "Perkiraan lokasi" in values
    assert "Terpetakan" in values


def test_burned_area_sheet_warns_that_hectares_are_not_comparable_to_hotspot_counts() -> None:
    """Dua metrik ini beda satuan dan beda kesegaran data (bulanan vs 3 jam).
    Tanpa peringatan tertulis, pembaca yang cuma membuka lampiran ini mudah
    menjumlahkannya dengan jumlah titik hotspot di sheet lain."""
    from app.services.export_service import BURNED_AREA_SHEET_TITLE, build_excel_file

    sheet = _load_sheet(
        build_excel_file([], burned_area_report=_BURNED_REPORT), BURNED_AREA_SHEET_TITLE
    )
    text = " ".join(
        str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None
    )

    assert "tidak dapat dijumlahkan langsung" in text
    assert "Kementerian Kehutanan" in text


def test_burned_area_sheet_handles_empty_report() -> None:
    from app.services.export_service import BURNED_AREA_SHEET_TITLE, build_excel_file

    empty = {"rows": [], "by_skema": [], "total_ha": 0.0, "kps_count": 0}
    sheet = _load_sheet(build_excel_file([], burned_area_report=empty), BURNED_AREA_SHEET_TITLE)
    text = " ".join(
        str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None
    )

    assert "Belum ada data luas terbakar" in text
